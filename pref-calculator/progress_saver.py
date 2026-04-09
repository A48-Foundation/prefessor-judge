"""Save and load pref session progress to/from Excel files.

Stores all session state needed to resume mid-workflow:
- Session metadata (state, rating scale, quota mode)
- All judges with their current scores
- Pairwise ranker state (Elo ratings, comparison history, queue)
- Matched/unmatched judge lists
"""
import io
import json

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# Styling constants
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="5865F2", end_color="5865F2", fill_type="solid")
META_FILL = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def _style_header(ws, row, col_count):
    """Apply header styling to a row."""
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER


def _score_display(score):
    """Convert internal score to display value."""
    if score is None:
        return ""
    if score >= 7.0:
        return "C"
    if score >= 6.0:
        return "S"
    return score


def save_progress(session, filename="prefs_progress.xlsx"):
    """Serialize a PrefSession to an Excel workbook and return bytes.

    Returns:
        tuple: (bytes, filename) — the Excel file content and suggested filename.
    """
    wb = openpyxl.Workbook()

    # --- Sheet 1: Metadata ---
    ws_meta = wb.active
    ws_meta.title = "Session Info"
    ws_meta.column_dimensions["A"].width = 25
    ws_meta.column_dimensions["B"].width = 50

    meta_fields = [
        ("State", session.state),
        ("Rating Max", session.rating_max),
        ("Current Index", session.current_idx),
        ("Quota Mode", session.quota_mode or ""),
        ("Quotas JSON", json.dumps(session.quotas) if session.quotas else ""),
        ("Ordinal Mode", str(getattr(session, "ordinal_mode", False))),
        ("Ordinal Rankings JSON", json.dumps(getattr(session, "ordinal_rankings", {}))),
        ("Ordinal Refine Tier", str(getattr(session, "ordinal_refine_tier", "") or "")),
    ]

    ws_meta.cell(row=1, column=1, value="Field")
    ws_meta.cell(row=1, column=2, value="Value")
    _style_header(ws_meta, 1, 2)

    for i, (field, value) in enumerate(meta_fields, start=2):
        ws_meta.cell(row=i, column=1, value=field).font = Font(bold=True)
        ws_meta.cell(row=i, column=2, value=str(value))

    # --- Sheet 2: All Judges with Scores ---
    ws_judges = wb.create_sheet("Judges")
    headers = ["Name", "School", "Rounds", "Score", "Source"]
    for col, h in enumerate(headers, 1):
        ws_judges.cell(row=1, column=col, value=h)
    _style_header(ws_judges, 1, len(headers))

    ws_judges.column_dimensions["A"].width = 30
    ws_judges.column_dimensions["B"].width = 25
    ws_judges.column_dimensions["C"].width = 10
    ws_judges.column_dimensions["D"].width = 10
    ws_judges.column_dimensions["E"].width = 20

    row = 2
    # Matched judges (from Notion)
    for csv_j, notion_name, notion_score in session.matched:
        name = csv_j["name"]
        final_score = session.scores_map.get(name, notion_score)
        ws_judges.cell(row=row, column=1, value=name)
        ws_judges.cell(row=row, column=2, value=csv_j.get("school", ""))
        ws_judges.cell(row=row, column=3, value=csv_j.get("rounds", 0))
        ws_judges.cell(row=row, column=4, value=_score_display(final_score))
        ws_judges.cell(row=row, column=5, value="Matched")
        row += 1

    # Unmatched judges
    for judge in session.unmatched:
        name = judge["name"]
        score = session.scores_map.get(name)
        ws_judges.cell(row=row, column=1, value=name)
        ws_judges.cell(row=row, column=2, value=judge.get("school", ""))
        ws_judges.cell(row=row, column=3, value=judge.get("rounds", 0))
        ws_judges.cell(row=row, column=4, value=_score_display(score))
        ws_judges.cell(row=row, column=5, value="Unmatched")
        row += 1

    # Pre-filled unmatched
    for judge in session.prefilled_unmatched:
        name = judge["name"]
        score = session.scores_map.get(name)
        ws_judges.cell(row=row, column=1, value=name)
        ws_judges.cell(row=row, column=2, value=judge.get("school", ""))
        ws_judges.cell(row=row, column=3, value=judge.get("rounds", 0))
        ws_judges.cell(row=row, column=4, value=_score_display(score))
        ws_judges.cell(row=row, column=5, value="Pre-filled")
        row += 1

    # Skipped judges
    for judge in session.skipped_judges:
        name = judge["name"]
        ws_judges.cell(row=row, column=1, value=name)
        ws_judges.cell(row=row, column=2, value=judge.get("school", ""))
        ws_judges.cell(row=row, column=3, value=judge.get("rounds", 0))
        ws_judges.cell(row=row, column=4, value="")
        ws_judges.cell(row=row, column=5, value="Skipped")
        row += 1

    # --- Sheet 3: Pairwise Ranker State (if active) ---
    if session.ranker:
        ws_elo = wb.create_sheet("Pairwise State")
        elo_headers = ["Name", "Elo Rating", "Type"]
        for col, h in enumerate(elo_headers, 1):
            ws_elo.cell(row=1, column=col, value=h)
        _style_header(ws_elo, 1, len(elo_headers))

        ws_elo.column_dimensions["A"].width = 30
        ws_elo.column_dimensions["B"].width = 15
        ws_elo.column_dimensions["C"].width = 15

        row = 2
        ranker = session.ranker
        for j in ranker.unknown:
            ws_elo.cell(row=row, column=1, value=j["name"])
            ws_elo.cell(row=row, column=2, value=round(ranker.elo.get(j["name"], 1500), 1))
            ws_elo.cell(row=row, column=3, value="Unknown")
            row += 1
        for j in ranker.anchors:
            ws_elo.cell(row=row, column=1, value=j["name"])
            ws_elo.cell(row=row, column=2, value=round(ranker.elo.get(j["name"], 1500), 1))
            ws_elo.cell(row=row, column=3, value="Anchor")
            row += 1

        # Comparison history
        ws_history = wb.create_sheet("Comparison History")
        hist_headers = ["Winner", "Loser"]
        for col, h in enumerate(hist_headers, 1):
            ws_history.cell(row=1, column=col, value=h)
        _style_header(ws_history, 1, len(hist_headers))

        ws_history.column_dimensions["A"].width = 30
        ws_history.column_dimensions["B"].width = 30

        for i, (winner, loser) in enumerate(ranker.history, start=2):
            ws_history.cell(row=i, column=1, value=winner)
            ws_history.cell(row=i, column=2, value=loser)

        # Remaining pairs queue
        ws_queue = wb.create_sheet("Remaining Pairs")
        queue_headers = ["Judge A", "Judge B"]
        for col, h in enumerate(queue_headers, 1):
            ws_queue.cell(row=1, column=col, value=h)
        _style_header(ws_queue, 1, len(queue_headers))

        ws_queue.column_dimensions["A"].width = 30
        ws_queue.column_dimensions["B"].width = 30

        for i, (a, b) in enumerate(ranker._pair_queue, start=2):
            ws_queue.cell(row=i, column=1, value=a["name"])
            ws_queue.cell(row=i, column=2, value=b["name"])

        # Ranker metadata
        ws_rmeta = wb.create_sheet("Ranker Meta")
        ws_rmeta.column_dimensions["A"].width = 25
        ws_rmeta.column_dimensions["B"].width = 50

        ws_rmeta.cell(row=1, column=1, value="Field")
        ws_rmeta.cell(row=1, column=2, value="Value")
        _style_header(ws_rmeta, 1, 2)

        ranker_meta = [
            ("Comparisons Done", ranker.comparisons_done),
            ("Current Round", ranker._current_round),
            ("Total Rounds", ranker._rounds_total),
            ("Matched Pairs JSON", json.dumps(list(ranker._matched_pairs))),
        ]
        for i, (field, value) in enumerate(ranker_meta, start=2):
            ws_rmeta.cell(row=i, column=1, value=field).font = Font(bold=True)
            ws_rmeta.cell(row=i, column=2, value=str(value))

    # --- Sheet: CSV Judges (original data for reconstruction) ---
    ws_csv = wb.create_sheet("CSV Judges")
    csv_headers = ["Name", "School", "Rounds", "Rating"]
    for col, h in enumerate(csv_headers, 1):
        ws_csv.cell(row=1, column=col, value=h)
    _style_header(ws_csv, 1, len(csv_headers))

    ws_csv.column_dimensions["A"].width = 30
    ws_csv.column_dimensions["B"].width = 25
    ws_csv.column_dimensions["C"].width = 10
    ws_csv.column_dimensions["D"].width = 10

    for i, j in enumerate(session.csv_judges, start=2):
        ws_csv.cell(row=i, column=1, value=j["name"])
        ws_csv.cell(row=i, column=2, value=j.get("school", ""))
        ws_csv.cell(row=i, column=3, value=j.get("rounds", 0))
        ws_csv.cell(row=i, column=4, value=j.get("rating", ""))

    # --- Sheet: Scores Map ---
    ws_scores = wb.create_sheet("Scores Map")
    sm_headers = ["Name", "Internal Score"]
    for col, h in enumerate(sm_headers, 1):
        ws_scores.cell(row=1, column=col, value=h)
    _style_header(ws_scores, 1, len(sm_headers))

    ws_scores.column_dimensions["A"].width = 30
    ws_scores.column_dimensions["B"].width = 15

    for i, (name, score) in enumerate(session.scores_map.items(), start=2):
        ws_scores.cell(row=i, column=1, value=name)
        ws_scores.cell(row=i, column=2, value=score)

    # --- Sheet: Matched Info ---
    ws_matched = wb.create_sheet("Matched Info")
    m_headers = ["CSV Name", "Notion Name", "Notion Score"]
    for col, h in enumerate(m_headers, 1):
        ws_matched.cell(row=1, column=col, value=h)
    _style_header(ws_matched, 1, len(m_headers))

    ws_matched.column_dimensions["A"].width = 30
    ws_matched.column_dimensions["B"].width = 30
    ws_matched.column_dimensions["C"].width = 15

    for i, (csv_j, notion_name, score) in enumerate(session.matched, start=2):
        ws_matched.cell(row=i, column=1, value=csv_j["name"])
        ws_matched.cell(row=i, column=2, value=notion_name or "")
        ws_matched.cell(row=i, column=3, value=score if score is not None else "")

    # Write to bytes
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue(), filename


def is_progress_file(file_bytes):
    """Check if an Excel file is a Prefessor Judge progress file.

    Returns True only if the workbook contains the expected 'Session Info' sheet
    with recognizable metadata fields.
    """
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        if "Session Info" not in wb.sheetnames:
            wb.close()
            return False
        ws = wb["Session Info"]
        fields = set()
        for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
            if row[0]:
                fields.add(str(row[0]))
        wb.close()
        return "State" in fields and "Rating Max" in fields
    except Exception:
        return False


def load_progress(file_bytes):
    """Deserialize an Excel progress file back into session data.

    Returns:
        dict with all the fields needed to reconstruct a PrefSession.
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    data = {}

    # --- Session Info ---
    ws_meta = wb["Session Info"]
    meta = {}
    for row in ws_meta.iter_rows(min_row=2, max_col=2, values_only=True):
        if row[0]:
            meta[row[0]] = row[1]
    data["state"] = meta.get("State", "awaiting_csv")
    data["rating_max"] = int(meta.get("Rating Max", 5))
    data["current_idx"] = int(meta.get("Current Index", 0))
    data["quota_mode"] = meta.get("Quota Mode") or None
    quotas_json = meta.get("Quotas JSON", "")
    if quotas_json:
        raw = json.loads(quotas_json)
        # JSON keys are strings — convert back to int
        data["quotas"] = {int(k): v for k, v in raw.items()}
    else:
        data["quotas"] = {}

    # Ordinal fields
    data["ordinal_mode"] = meta.get("Ordinal Mode", "False") == "True"
    ordinal_json = meta.get("Ordinal Rankings JSON", "")
    if ordinal_json:
        raw_ord = json.loads(ordinal_json)
        data["ordinal_rankings"] = {int(k): v for k, v in raw_ord.items()}
    else:
        data["ordinal_rankings"] = {}
    refine_tier = meta.get("Ordinal Refine Tier", "")
    data["ordinal_refine_tier"] = int(refine_tier) if refine_tier and refine_tier.isdigit() else None

    # --- CSV Judges ---
    ws_csv = wb["CSV Judges"]
    csv_judges = []
    for row in ws_csv.iter_rows(min_row=2, max_col=4, values_only=True):
        if not row[0]:
            continue
        csv_judges.append({
            "name": str(row[0]),
            "school": str(row[1] or ""),
            "rounds": int(row[2] or 0),
            "rating": str(row[3] or ""),
        })
    data["csv_judges"] = csv_judges

    # --- Scores Map ---
    ws_scores = wb["Scores Map"]
    scores_map = {}
    for row in ws_scores.iter_rows(min_row=2, max_col=2, values_only=True):
        if row[0] is not None and row[1] is not None:
            scores_map[str(row[0])] = float(row[1])
    data["scores_map"] = scores_map

    # --- Matched Info ---
    ws_matched = wb["Matched Info"]
    matched_raw = []
    for row in ws_matched.iter_rows(min_row=2, max_col=3, values_only=True):
        if not row[0]:
            continue
        matched_raw.append({
            "csv_name": str(row[0]),
            "notion_name": str(row[1]) if row[1] else None,
            "notion_score": float(row[2]) if row[2] is not None and row[2] != "" else None,
        })
    data["matched_raw"] = matched_raw

    # --- Judges sheet (for source categories) ---
    ws_judges = wb["Judges"]
    unmatched_names = set()
    prefilled_names = set()
    skipped_names = set()
    for row in ws_judges.iter_rows(min_row=2, max_col=5, values_only=True):
        if not row[0]:
            continue
        source = str(row[4] or "")
        name = str(row[0])
        if source == "Unmatched":
            unmatched_names.add(name)
        elif source == "Pre-filled":
            prefilled_names.add(name)
        elif source == "Skipped":
            skipped_names.add(name)
    data["unmatched_names"] = unmatched_names
    data["prefilled_names"] = prefilled_names
    data["skipped_names"] = skipped_names

    # --- Pairwise State (optional) ---
    data["has_ranker"] = "Pairwise State" in wb.sheetnames
    if data["has_ranker"]:
        ws_elo = wb["Pairwise State"]
        elo_data = {}
        unknown_names = []
        anchor_names = []
        for row in ws_elo.iter_rows(min_row=2, max_col=3, values_only=True):
            if not row[0]:
                continue
            name = str(row[0])
            elo_data[name] = float(row[1] or 1500)
            if str(row[2]) == "Unknown":
                unknown_names.append(name)
            else:
                anchor_names.append(name)
        data["elo_data"] = elo_data
        data["unknown_names"] = unknown_names
        data["anchor_names"] = anchor_names

        # History
        ws_history = wb["Comparison History"]
        history = []
        for row in ws_history.iter_rows(min_row=2, max_col=2, values_only=True):
            if row[0]:
                history.append((str(row[0]), str(row[1])))
        data["history"] = history

        # Remaining pairs
        ws_queue = wb["Remaining Pairs"]
        pair_queue = []
        for row in ws_queue.iter_rows(min_row=2, max_col=2, values_only=True):
            if row[0]:
                pair_queue.append((str(row[0]), str(row[1])))
        data["pair_queue"] = pair_queue

        # Ranker metadata
        ws_rmeta = wb["Ranker Meta"]
        rmeta = {}
        for row in ws_rmeta.iter_rows(min_row=2, max_col=2, values_only=True):
            if row[0]:
                rmeta[str(row[0])] = row[1]
        data["comparisons_done"] = int(rmeta.get("Comparisons Done", 0))
        data["current_round"] = int(rmeta.get("Current Round", 0))
        data["total_rounds"] = int(rmeta.get("Total Rounds", 6))
        mp_json = rmeta.get("Matched Pairs JSON", "[]")
        data["matched_pairs"] = set(tuple(p) for p in json.loads(mp_json))

    wb.close()
    return data


def restore_session(session, data):
    """Apply loaded progress data onto a PrefSession object.

    Reconstructs csv_judges, matched, unmatched, scores_map, and
    optionally the PairwiseRanker with its full Elo/queue state.
    """
    from pairwise_ranker import PairwiseRanker

    session.state = data["state"]
    session.rating_max = data["rating_max"]
    session.current_idx = data["current_idx"]
    session.quota_mode = data["quota_mode"]
    session.quotas = data["quotas"]
    session.csv_judges = data["csv_judges"]
    session.scores_map = data["scores_map"]

    # Ordinal fields
    session.ordinal_mode = data.get("ordinal_mode", False)
    session.ordinal_rankings = data.get("ordinal_rankings", {})
    session.ordinal_refine_tier = data.get("ordinal_refine_tier")

    # Build name→judge lookup from csv_judges
    judge_lookup = {j["name"]: j for j in session.csv_judges}

    # Reconstruct matched list: [(csv_judge_dict, notion_name, notion_score), ...]
    session.matched = []
    for m in data["matched_raw"]:
        csv_j = judge_lookup.get(m["csv_name"])
        if csv_j:
            session.matched.append((csv_j, m["notion_name"], m["notion_score"]))

    # Reconstruct unmatched, prefilled_unmatched, skipped
    session.unmatched = [j for j in session.csv_judges if j["name"] in data["unmatched_names"]]
    session.prefilled_unmatched = [j for j in session.csv_judges if j["name"] in data["prefilled_names"]]
    session.skipped_judges = [j for j in session.csv_judges if j["name"] in data["skipped_names"]]

    # Reconstruct pairwise ranker if it was active
    if data.get("has_ranker"):
        unknown_judges = [judge_lookup[n] for n in data["unknown_names"] if n in judge_lookup]
        anchor_judges = []
        for n in data["anchor_names"]:
            j = judge_lookup.get(n)
            if j:
                aj = dict(j)
                aj["score"] = data["scores_map"].get(n)
                anchor_judges.append(aj)

        # Create ranker but override its auto-generated state
        ranker = PairwiseRanker.__new__(PairwiseRanker)
        ranker.unknown = unknown_judges
        ranker.anchors = anchor_judges
        ranker.all_judges = unknown_judges + anchor_judges
        ranker.elo = data["elo_data"]
        ranker.comparisons_done = data["comparisons_done"]
        ranker.history = data["history"]
        ranker._undo_stack = []
        ranker._matched_pairs = data["matched_pairs"]
        ranker._current_round = data["current_round"]
        ranker._rounds_total = data["total_rounds"]

        # Rebuild pair queue from saved names
        ranker._pair_queue = []
        for a_name, b_name in data["pair_queue"]:
            a_judge = judge_lookup.get(a_name)
            b_judge = judge_lookup.get(b_name)
            # Anchors aren't in judge_lookup by default, check anchor list too
            if not a_judge:
                a_judge = next((aj for aj in anchor_judges if aj["name"] == a_name), None)
            if not b_judge:
                b_judge = next((aj for aj in anchor_judges if aj["name"] == b_name), None)
            if a_judge and b_judge:
                ranker._pair_queue.append((a_judge, b_judge))

        session.ranker = ranker
