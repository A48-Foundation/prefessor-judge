"""Prefessor Judge - Discord Bot.

A Discord bot that runs the pref sheet generator through chat.
Interactive UI with buttons, embeds, judge comparison, and review/edit.

Flow: user uploads CSV → bot matches judges → choose scoring method
(manual / skip / pairwise compare) → interactive scoring with compare
feature → review & edit → quota mode → tier quotas → output CSV.
"""
import csv as csv_mod
import io
import os
import sys
import tempfile
import urllib.parse

import discord
from discord import ui
from dotenv import load_dotenv

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "pref-calculator"))
from csv_parser import parse_tournament_csv
from notion_reader import fetch_notion_judges
from name_matcher import match_judges
from tier_assigner import assign_tiers, format_report
from csv_writer import write_output_csv
from pairwise_ranker import PairwiseRanker
from progress_saver import save_progress, load_progress, restore_session, is_progress_file

try:
    from judge_scraper import TabroomScraper
    from tabroom_cache import TabroomCache
except ImportError:
    TabroomScraper = None
    TabroomCache = None

load_dotenv()

intents = discord.Intents.default()
intents.message_content = True
client = discord.Client(intents=intents)

# Shared Tabroom scraper and cache (initialized on ready)
tabroom_scraper = None
tabroom_cache = TabroomCache() if TabroomCache else None

# Per-channel session state
sessions: dict[int, "PrefSession"] = {}

EMBED_COLOR = 0x5865F2
WARN_COLOR = 0xFEE75C
SUCCESS_COLOR = 0x57F287
ERROR_COLOR = 0xED4245


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def tabroom_paradigm_url(name: str) -> str:
    if ", " in name:
        last, first = name.split(", ", 1)
    else:
        parts = name.split()
        first = parts[0] if parts else ""
        last = " ".join(parts[1:]) if len(parts) > 1 else ""
    params = urllib.parse.urlencode({"search_first": first.strip(), "search_last": last.strip()})
    return f"https://www.tabroom.com/index/paradigm.mhtml?{params}"


def truncate(text: str, limit: int = 500) -> str:
    if not text:
        return "*No paradigm available*"
    if len(text) <= limit:
        return text
    return text[:limit].rsplit(" ", 1)[0] + "…"


def parse_all_quotas(text: str, mode: str) -> dict[int, dict] | None:
    """Parse a single-line quota input for all 6 tiers.

    Formats accepted:
      6 8 10 8 6 -        (space-separated, one value per tier = min only)
      6,8 8,10 10,- 8,- 6,- -   (min,max per tier)
    Use '-' or 'skip' to skip a tier.
    """
    text = text.strip()
    if not text:
        return None

    # Split by whitespace
    parts = text.split()
    if len(parts) < 5 or len(parts) > 12:
        return None

    # If parts look like "6,8" they have min,max; otherwise just min
    quotas = {}
    tier_values = []

    # Try to collect 6 tier entries
    for p in parts:
        p = p.strip().rstrip(",")
        if p in ("-", "skip", "0"):
            tier_values.append(None)
        elif "," in p:
            tier_values.append(p)
        else:
            tier_values.append(p)

    if len(tier_values) < 5:
        return None
    # Pad to 6 if only 5 given (strike tier optional)
    while len(tier_values) < 6:
        tier_values.append(None)

    for i, val in enumerate(tier_values[:6]):
        tier = i + 1
        if val is None:
            continue
        q: dict[str, int] = {}
        if "," in val:
            sub = val.split(",", 1)
            if sub[0].strip() not in ("-", ""):
                q["min"] = int(sub[0].strip())
            if sub[1].strip() not in ("-", ""):
                q["max"] = int(sub[1].strip())
        else:
            q["min"] = int(val)
        if q:
            quotas[tier] = q

    return quotas if quotas else None


def _score_label(score: float) -> str:
    """Return display label for an internal score: 'C' for conflict, 'S' for strike, else the number."""
    if score >= 7.0:
        return "C"
    if score >= 6.0:
        return "S"
    return str(score)


def build_judge_embed(judge: dict, paradigm: dict | None, *, score: float | None = None,
                      index: int | None = None, total: int | None = None) -> discord.Embed:
    title = judge["name"]
    if index is not None and total is not None:
        title = f"Judge {index}/{total} — {judge['name']}"
    embed = discord.Embed(title=title, color=EMBED_COLOR)
    embed.add_field(name="🏫 School", value=judge.get("school") or "Unknown", inline=True)
    embed.add_field(name="🔄 Rounds", value=str(judge.get("rounds", "?")), inline=True)
    if score is not None:
        label = _score_label(score)
        embed.add_field(name="⭐ Score", value=label, inline=True)
    url = tabroom_paradigm_url(judge["name"])
    if paradigm and paradigm.get("philosophy"):
        embed.add_field(name="📖 Paradigm", value=truncate(paradigm["philosophy"]), inline=False)
    embed.add_field(name="🔗 Tabroom", value=f"[View full paradigm]({url})", inline=False)
    if index is not None and total is not None:
        embed.set_footer(text=f"Progress: {index}/{total}")
    return embed


def split_name(name: str) -> tuple[str, str]:
    if ", " in name:
        last, first = name.split(", ", 1)
        return first.strip(), last.strip()
    parts = name.split()
    first = parts[0] if parts else ""
    last = " ".join(parts[1:]) if len(parts) > 1 else ""
    return first, last


def _normalize_score(raw: float, rating_max: int) -> float:
    """Normalize a raw score from [1, rating_max] to internal [1.0, 5.0].

    Values >= rating_max + 2 map to 7.0 (conflict).
    Values >= rating_max + 1 map to 6.0 (strike).
    For rating_max == 5 this is an identity.
    """
    if raw >= rating_max + 2:
        return 7.0
    if raw >= rating_max + 1:
        return 6.0
    if rating_max == 5:
        return raw
    return 1.0 + (raw - 1) * 4.0 / (rating_max - 1)


def _parse_prefilled_rating(rating_str: str, rating_max: int = 5) -> float | None:
    """Convert a CSV rating string to an internal score, or None if empty/invalid.

    Values 1..rating_max are normalized to the internal 1-5 scale.
    rating_max+1 and 'S' map to 6.0 (strike).
    'C' maps to 7.0 (conflict).
    """
    s = rating_str.strip().upper()
    if not s:
        return None
    if s == "S":
        return 6.0
    if s == "C":
        return 7.0
    try:
        val = float(s)
        if val < 1:
            return None
        internal = _normalize_score(val, rating_max)
        return round(internal * 2) / 2  # snap to 0.5
    except ValueError:
        return None


def _process_csv_prefilled(session: "PrefSession") -> int:
    """Extract pre-filled ratings from CSV judges into scores_map.

    Returns count of pre-filled judges.
    """
    count = 0
    for judge in session.csv_judges:
        score = _parse_prefilled_rating(judge.get("rating", ""),
                                        session.rating_max)
        if score is not None:
            session.scores_map[judge["name"]] = score
            count += 1
    return count


def _select_anchors(matched, count=4):
    """Pick a spread of anchor judges from matched list for Elo calibration."""
    if not matched:
        return []
    scored = [(csv_j, name, score) for csv_j, name, score in matched if score is not None]
    if not scored:
        return []
    scored.sort(key=lambda x: x[2])
    step = max(1, len(scored) // count)
    selected = scored[::step][:count]
    return [{"name": csv_j["name"], "school": csv_j["school"],
             "rounds": csv_j["rounds"], "score": score}
            for csv_j, _, score in selected]


# ---------------------------------------------------------------------------
# Session
# ---------------------------------------------------------------------------

class PrefSession:
    def __init__(self, channel_id: int, user_id: int):
        self.channel_id = channel_id
        self.user_id = user_id
        self.state = "awaiting_csv"
        self.csv_judges: list[dict] = []
        self.matched: list[tuple] = []
        self.unmatched: list[dict] = []
        self.skipped_judges: list[dict] = []
        self.scores_map: dict[str, float] = {}
        self.rating_max: int = 5
        self.current_idx: int = 0
        self.quota_mode: str | None = None
        self.quotas: dict[int, dict] = {}
        self.current_tier: int = 1
        self.notion_judges: dict = {}
        self.paradigms: dict[str, dict | None] = {}
        self.ranker: PairwiseRanker | None = None
        self.paradigm_messages: list = []  # messages to delete on next comparison
        self.prefilled_unmatched: list[dict] = []  # pre-rated judges split from unmatched
        # Ordinal ranking mode
        self.ordinal_mode: bool = False
        self.ordinal_rankings: dict[int, list[str]] = {}  # tier → [judge names in order]
        self.ordinal_refine_tier: int | None = None  # which tier is being refined


# ---------------------------------------------------------------------------
# Rating Range View
# ---------------------------------------------------------------------------

class RatingRangeView(ui.View):
    """Let user choose the tournament's rating scale (e.g. 1-5, 1-6, 1-7)."""

    def __init__(self, session: PrefSession):
        super().__init__(timeout=None)
        self.session = session

    @ui.button(label="1–5", style=discord.ButtonStyle.primary)
    async def range_5(self, interaction: discord.Interaction, button: ui.Button):
        await self._set_range(interaction, 5)

    @ui.button(label="1–6", style=discord.ButtonStyle.primary)
    async def range_6(self, interaction: discord.Interaction, button: ui.Button):
        await self._set_range(interaction, 6)

    @ui.button(label="1–7", style=discord.ButtonStyle.primary)
    async def range_7(self, interaction: discord.Interaction, button: ui.Button):
        await self._set_range(interaction, 7)

    async def _set_range(self, interaction: discord.Interaction, max_val: int):
        self.stop()
        await _apply_rating_range(interaction, self.session, max_val)


# ---------------------------------------------------------------------------
# Interactive Views — Scoring
# ---------------------------------------------------------------------------

class ScoreButton(ui.Button):
    def __init__(self, raw_score: float, label: str, style: discord.ButtonStyle):
        super().__init__(label=label, style=style, custom_id=f"score_{raw_score}")
        self.raw_score = raw_score

    async def callback(self, interaction: discord.Interaction):
        view: ScoringView = self.view  # type: ignore
        session = sessions.get(interaction.channel_id)
        if not session or interaction.user.id != session.user_id:
            await interaction.response.send_message("Not your session.", ephemeral=True)
            return
        judge = session.unmatched[session.current_idx]
        session.scores_map[judge["name"]] = _normalize_score(self.raw_score, session.rating_max)
        session.current_idx += 1
        if session.current_idx < len(session.unmatched):
            await view.show_judge(interaction)
        else:
            await view.finish_scoring(interaction)


class ScoringView(ui.View):
    def __init__(self, session: PrefSession):
        super().__init__(timeout=None)
        self.session = session
        rm = session.rating_max
        for s in range(1, rm + 1):
            if s <= rm * 0.4:
                style = discord.ButtonStyle.success
            elif s <= rm * 0.8:
                style = discord.ButtonStyle.primary
            else:
                style = discord.ButtonStyle.secondary
            self.add_item(ScoreButton(float(s), str(s), style))
        # Strike button
        self.add_item(ScoreButton(float(rm + 1), "Strike",
                                  discord.ButtonStyle.danger))
        # Conflict button
        self.add_item(ScoreButton(float(rm + 2), "Conflict",
                                  discord.ButtonStyle.secondary))

    @ui.button(label="◀ Previous", style=discord.ButtonStyle.secondary, row=2)
    async def prev_button(self, interaction: discord.Interaction, button: ui.Button):
        if self.session.current_idx > 0:
            self.session.current_idx -= 1
        await self.show_judge(interaction)

    @ui.button(label="⏭ Skip", style=discord.ButtonStyle.secondary, row=2)
    async def skip_button(self, interaction: discord.Interaction, button: ui.Button):
        judge = self.session.unmatched[self.session.current_idx]
        self.session.scores_map[judge["name"]] = 4.0  # internal: below average
        self.session.current_idx += 1
        if self.session.current_idx < len(self.session.unmatched):
            await self.show_judge(interaction)
        else:
            await self.finish_scoring(interaction)

    @ui.button(label="💾 Save", style=discord.ButtonStyle.secondary, row=2)
    async def save_button(self, interaction: discord.Interaction, button: ui.Button):
        session = self.session
        file_bytes, filename = save_progress(session)
        file = discord.File(io.BytesIO(file_bytes), filename=filename)
        await interaction.response.send_message(
            embed=discord.Embed(
                title="💾 Progress Saved",
                description="Download this file to resume later.\n"
                            "Upload it when starting a new session to pick up where you left off.",
                color=SUCCESS_COLOR),
            file=file, ephemeral=True)

    @ui.button(label="🔍 Compare", style=discord.ButtonStyle.primary, row=2)
    async def compare_button(self, interaction: discord.Interaction, button: ui.Button):
        current_judge = self.session.unmatched[self.session.current_idx]
        all_names = [j["name"] for j in self.session.csv_judges if j["name"] != current_judge["name"]]
        options = [discord.SelectOption(label=n[:100], value=n[:100]) for n in all_names[:25]]
        if not options:
            await interaction.response.send_message("No other judges to compare.", ephemeral=True)
            return
        view = CompareSelectView(self.session, current_judge, options, parent_view=self)
        await interaction.response.send_message("Pick a judge to compare with:", view=view, ephemeral=True)

    async def show_judge(self, interaction: discord.Interaction):
        session = self.session
        judge = session.unmatched[session.current_idx]
        idx = session.current_idx + 1
        total = len(session.unmatched)
        existing_score = session.scores_map.get(judge["name"])
        # Build compact info embed (no paradigm text — that goes in separate message)
        embed = discord.Embed(title=f"Judge {idx}/{total} — {judge['name']}", color=EMBED_COLOR)
        embed.add_field(name="🏫 School", value=judge.get("school") or "Unknown", inline=True)
        embed.add_field(name="🔄 Rounds", value=str(judge.get("rounds", "?")), inline=True)
        if existing_score is not None:
            label = _score_label(existing_score)
            embed.add_field(name="⭐ Score", value=label, inline=True)
        url = tabroom_paradigm_url(judge["name"])
        embed.add_field(name="🔗 Tabroom", value=f"[View full paradigm]({url})", inline=False)
        embed.set_footer(text=f"Progress: {idx}/{total}" +
                         (f" • Previously scored: {existing_score}" if existing_score is not None else ""))
        await interaction.response.edit_message(embed=embed, view=self)
        # Update paradigm as separate message below the buttons
        para_embed = _build_paradigm_embed(judge, session)
        await _update_paradigms(interaction.channel, session, [para_embed])

    async def finish_scoring(self, interaction: discord.Interaction):
        self.stop()
        # Clean up paradigm messages
        for msg in self.session.paradigm_messages:
            try:
                await msg.delete()
            except Exception:
                pass
        self.session.paradigm_messages.clear()
        self.session.state = "reviewing"
        review_view = ReviewView(self.session)
        embed = review_view.build_summary_embed()
        await interaction.response.edit_message(embed=embed, view=review_view)


class CompareSelectView(ui.View):
    def __init__(self, session: PrefSession, current_judge: dict,
                 options: list[discord.SelectOption], parent_view: ScoringView):
        super().__init__(timeout=None)
        self.session = session
        self.current_judge = current_judge
        self.parent_view = parent_view
        select = ui.Select(placeholder="Select a judge to compare…", options=options,
                           custom_id="compare_select")
        select.callback = self.on_select
        self.add_item(select)

    async def on_select(self, interaction: discord.Interaction):
        selected_name = interaction.data["values"][0]  # type: ignore
        other = next((j for j in self.session.csv_judges if j["name"] == selected_name), None)
        if not other:
            await interaction.response.send_message("Judge not found.", ephemeral=True)
            return
        p1 = self.session.paradigms.get(self.current_judge["name"])
        if p1 is None and tabroom_scraper and tabroom_cache:
            p1 = tabroom_cache.get_or_fetch(self.current_judge["name"], tabroom_scraper)
            self.session.paradigms[self.current_judge["name"]] = p1
        p2 = self.session.paradigms.get(other["name"])
        if p2 is None and tabroom_scraper and tabroom_cache:
            p2 = tabroom_cache.get_or_fetch(other["name"], tabroom_scraper)
            self.session.paradigms[other["name"]] = p2
        score1 = self.session.scores_map.get(self.current_judge["name"])
        score2 = self.session.scores_map.get(other["name"])
        if score2 is None:
            for csv_j, _, s in self.session.matched:
                if csv_j["name"] == other["name"]:
                    score2 = s
                    break
        embed1 = build_judge_embed(self.current_judge, p1, score=score1)
        embed1.title = f"🅰️ {self.current_judge['name']}"
        embed1.color = 0x5865F2
        embed2 = build_judge_embed(other, p2, score=score2)
        embed2.title = f"🅱️ {other['name']}"
        embed2.color = 0xEB459E
        back_view = BackToScoringView(self.parent_view)
        await interaction.response.edit_message(
            content="**Side-by-side comparison:**", embeds=[embed1, embed2], view=back_view)


class BackToScoringView(ui.View):
    def __init__(self, parent_view: ScoringView):
        super().__init__(timeout=None)
        self.parent_view = parent_view

    @ui.button(label="◀ Back to Scoring", style=discord.ButtonStyle.primary)
    async def back(self, interaction: discord.Interaction, button: ui.Button):
        session = self.parent_view.session
        judge = session.unmatched[session.current_idx]
        paradigm = session.paradigms.get(judge["name"])
        idx = session.current_idx + 1
        total = len(session.unmatched)
        embed = build_judge_embed(judge, paradigm, score=session.scores_map.get(judge["name"]),
                                  index=idx, total=total)
        await interaction.response.edit_message(content=None, embeds=[embed], view=self.parent_view)


# ---------------------------------------------------------------------------
# Review View
# ---------------------------------------------------------------------------

class ReviewEditSelect(ui.Select):
    def __init__(self, session: PrefSession):
        self.session = session
        options = []
        for j in session.unmatched:
            score = session.scores_map.get(j["name"], 4.0)
            label = f"{j['name']} — Score: {score}"
            options.append(discord.SelectOption(label=label[:100], value=j["name"][:100]))
        super().__init__(placeholder="Select a judge to re-score…",
                         options=options[:25], custom_id="review_select")

    async def callback(self, interaction: discord.Interaction):
        selected = interaction.data["values"][0]  # type: ignore
        judge = next((j for j in self.session.unmatched if j["name"] == selected), None)
        if not judge:
            await interaction.response.send_message("Judge not found.", ephemeral=True)
            return
        modal = ReScoreModal(self.session, judge)
        await interaction.response.send_modal(modal)


class ReScoreModal(ui.Modal, title="Re-score Judge"):
    score_input = ui.TextInput(label="Score", placeholder="e.g. 3, S for strike, C for conflict",
                               max_length=4, required=True)

    def __init__(self, session: PrefSession, judge: dict):
        super().__init__()
        self.session = session
        self.judge = judge
        current = session.scores_map.get(judge["name"], 4.0)
        self.score_input.default = str(current)
        rm = session.rating_max
        self.score_input.label = f"Score (1-{rm}, S=strike, C=conflict)"

    async def on_submit(self, interaction: discord.Interaction):
        rm = self.session.rating_max
        raw_text = self.score_input.value.strip().upper()
        if raw_text == "S":
            score = 6.0
        elif raw_text == "C":
            score = 7.0
        else:
            try:
                raw = float(raw_text)
                if raw < 1 or raw > rm + 1:
                    await interaction.response.send_message(
                        f"Score must be 1-{rm} (or {rm + 1} for strike).", ephemeral=True)
                    return
                score = _normalize_score(round(raw * 2) / 2, rm)
            except ValueError:
                await interaction.response.send_message("Invalid number.", ephemeral=True)
                return
        self.session.scores_map[self.judge["name"]] = score
        view: ReviewView = self.session._review_view  # type: ignore
        embed = view.build_summary_embed()
        await interaction.response.edit_message(embed=embed, view=view)


class ReviewView(ui.View):
    def __init__(self, session: PrefSession):
        super().__init__(timeout=None)
        self.session = session
        session._review_view = self
        if session.unmatched:
            self.add_item(ReviewEditSelect(session))

    def build_summary_embed(self) -> discord.Embed:
        embed = discord.Embed(title="📋 Score Review", color=SUCCESS_COLOR,
                              description="Review your scores. Use the dropdown to change any score, "
                                          "then press **Confirm & Continue**.")
        lines = []
        for j in self.session.unmatched:
            score = self.session.scores_map.get(j["name"], 4.0)
            label = _score_label(score)
            lines.append(f"• **{j['name']}** ({j['school']}) — {label}")
        embed.add_field(name="Scored Judges", value="\n".join(lines) or "None", inline=False)
        return embed

    @ui.button(label="✅ Confirm & Continue", style=discord.ButtonStyle.success, row=2)
    async def confirm(self, interaction: discord.Interaction, button: ui.Button):
        self.stop()
        self.session.state = "awaiting_quota_mode"
        await send_quota_mode_prompt(interaction.channel, self.session, interaction=interaction)

    @ui.button(label="💾 Save Progress", style=discord.ButtonStyle.secondary, row=2)
    async def save_review(self, interaction: discord.Interaction, button: ui.Button):
        session = self.session
        file_bytes, filename = save_progress(session)
        file = discord.File(io.BytesIO(file_bytes), filename=filename)
        await interaction.response.send_message(
            embed=discord.Embed(
                title="💾 Progress Saved",
                description="Download this file to resume later.\n"
                            "Upload it when starting a new session to pick up where you left off.",
                color=SUCCESS_COLOR),
            file=file, ephemeral=True)


# ---------------------------------------------------------------------------
# Pairwise Comparison Views
# ---------------------------------------------------------------------------

class RateSelect(ui.Select):
    """Dropdown to directly rate Judge A or B on the session's rating scale."""
    def __init__(self, which: str, session: PrefSession):
        self.which = which  # "a" or "b"
        self.session = session
        rm = session.rating_max
        side = "🅰️ Judge A" if which == "a" else "🅱️ Judge B"
        emojis = {1: "⭐", rm: "⛔"}
        options = []
        for s in range(1, rm + 1):
            if s == 1:
                lbl = f"{s} — Best"
            elif s == rm:
                lbl = f"{s} — Worst"
            elif s <= rm // 2:
                lbl = f"{s} — Good"
            elif s == (rm + 1) // 2:
                lbl = f"{s} — Average"
            else:
                lbl = f"{s} — Below Average"
            options.append(discord.SelectOption(
                label=lbl, value=str(s), emoji=emojis.get(s, "➖")))
        options.append(discord.SelectOption(
            label="Strike", value=str(rm + 1), emoji="🚫"))
        options.append(discord.SelectOption(
            label="Conflict", value=str(rm + 2), emoji="⚠️"))
        row = 2 if which == "a" else 3
        super().__init__(placeholder=f"Rate {side} directly…", options=options, row=row, min_values=1, max_values=1)

    async def callback(self, interaction: discord.Interaction):
        session = self.session
        if not session.ranker:
            return
        pair = session.ranker.next_pair()
        if not pair:
            return
        judge = pair[0] if self.which == "a" else pair[1]
        raw = float(self.values[0])
        rm = session.rating_max
        if raw >= rm + 2:
            internal = 7.0
            label = "Conflict"
        elif raw >= rm + 1:
            internal = 6.0
            label = "Strike"
        else:
            internal = _normalize_score(raw, rm)
            label = str(int(raw))
        session.scores_map[judge["name"]] = internal

        if internal >= 6.0:
            session.ranker.remove_judge(judge, internal)
        else:
            session.ranker.rate_judge(judge, internal)

        if session.ranker.is_complete:
            self.view.stop()
            await interaction.response.edit_message(content="⏳ Finishing comparison…", embeds=[], view=None)
            await _finish_comparison(interaction.channel, session)
        else:
            info_embeds, para_embeds = _build_comparison_embeds(session)
            await interaction.response.edit_message(
                content=f"✅ **{judge['name']}** rated **{label}**.",
                embeds=info_embeds, view=self.view)
            await _update_paradigms(interaction.channel, session, para_embeds)


class PairwiseView(ui.View):
    def __init__(self, session: PrefSession):
        super().__init__(timeout=None)
        self.session = session
        self.add_item(RateSelect("a", session))
        self.add_item(RateSelect("b", session))

    @ui.button(label="A", style=discord.ButtonStyle.primary, row=0)
    async def pick_a(self, interaction: discord.Interaction, button: ui.Button):
        await self._record(interaction, "a")

    @ui.button(label="B", style=discord.ButtonStyle.success, row=0)
    async def pick_b(self, interaction: discord.Interaction, button: ui.Button):
        await self._record(interaction, "b")

    @ui.button(label="Skip", style=discord.ButtonStyle.secondary, row=0)
    async def skip(self, interaction: discord.Interaction, button: ui.Button):
        await self._record(interaction, "skip")

    @ui.button(label="Undo", style=discord.ButtonStyle.secondary, row=0)
    async def undo(self, interaction: discord.Interaction, button: ui.Button):
        session = self.session
        if not session.ranker:
            return
        # Undo strike/conflict score if the last action was a special removal
        if session.ranker._undo_stack and session.ranker._undo_stack[-1]["type"] == "special":
            last = session.ranker._undo_stack[-1]
            judge_name = last["judge"]["name"]
            session.scores_map.pop(judge_name, None)
        if session.ranker.undo():
            info_embeds, para_embeds = _build_comparison_embeds(session)
            await interaction.response.edit_message(content="↩️ Undone.", embeds=info_embeds, view=self)
            await _update_paradigms(interaction.channel, session, para_embeds)
        else:
            await interaction.response.edit_message(content="Nothing to undo.")

    @ui.button(label="Done", style=discord.ButtonStyle.secondary, row=0)
    async def done(self, interaction: discord.Interaction, button: ui.Button):
        self.stop()
        await interaction.response.edit_message(content="⏳ Finishing comparison…", embeds=[], view=None)
        await _finish_comparison(interaction.channel, self.session)

    @ui.button(label="💾 Save", style=discord.ButtonStyle.secondary, row=1)
    async def save_progress_btn(self, interaction: discord.Interaction, button: ui.Button):
        session = self.session
        file_bytes, filename = save_progress(session)
        file = discord.File(io.BytesIO(file_bytes), filename=filename)
        await interaction.response.send_message(
            embed=discord.Embed(
                title="💾 Progress Saved",
                description="Download this file to resume later.\n"
                            "Upload it when starting a new session to pick up where you left off.",
                color=SUCCESS_COLOR),
            file=file, ephemeral=True)

    @ui.button(label="Strike A", style=discord.ButtonStyle.danger, row=1)
    async def strike_a(self, interaction: discord.Interaction, button: ui.Button):
        await self._assign_special(interaction, "a", 6.0)

    @ui.button(label="Strike B", style=discord.ButtonStyle.danger, row=1)
    async def strike_b(self, interaction: discord.Interaction, button: ui.Button):
        await self._assign_special(interaction, "b", 6.0)

    @ui.button(label="Conflict A", style=discord.ButtonStyle.secondary, row=1)
    async def conflict_a(self, interaction: discord.Interaction, button: ui.Button):
        await self._assign_special(interaction, "a", 7.0)

    @ui.button(label="Conflict B", style=discord.ButtonStyle.secondary, row=1)
    async def conflict_b(self, interaction: discord.Interaction, button: ui.Button):
        await self._assign_special(interaction, "b", 7.0)

    async def _assign_special(self, interaction: discord.Interaction, which: str,
                              score: float = 6.0):
        """Assign strike (6.0) or conflict (7.0) to a judge, removing them from comparisons."""
        session = self.session
        if not session.ranker:
            return
        pair = session.ranker.next_pair()
        if not pair:
            self.stop()
            await interaction.response.edit_message(content="⏳ Finishing comparison…", embeds=[], view=None)
            await _finish_comparison(interaction.channel, session)
            return
        judge = pair[0] if which == "a" else pair[1]
        label = "Conflict" if score == 7.0 else "Strike"
        session.scores_map[judge["name"]] = score
        session.ranker.remove_judge(judge, score)
        if session.ranker.is_complete:
            self.stop()
            await interaction.response.edit_message(content="⏳ Finishing comparison…", embeds=[], view=None)
            await _finish_comparison(interaction.channel, session)
        else:
            info_embeds, para_embeds = _build_comparison_embeds(session)
            await interaction.response.edit_message(
                content=f"✅ **{judge['name']}** marked as **{label}**.",
                embeds=info_embeds, view=self)
            await _update_paradigms(interaction.channel, session, para_embeds)

    async def _record(self, interaction: discord.Interaction, choice: str):
        session = self.session
        if not session.ranker:
            return
        pair = session.ranker.next_pair()
        if not pair:
            self.stop()
            await interaction.response.edit_message(content="⏳ Finishing comparison…", embeds=[], view=None)
            await _finish_comparison(interaction.channel, session)
            return
        judge_a, judge_b = pair
        if choice == "a":
            session.ranker.record_result(judge_a, judge_b)
        elif choice == "b":
            session.ranker.record_result(judge_b, judge_a)
        else:
            session.ranker.skip_pair()
        if session.ranker.is_complete:
            self.stop()
            await interaction.response.edit_message(content="⏳ Finishing comparison…", embeds=[], view=None)
            await _finish_comparison(interaction.channel, session)
        else:
            info_embeds, para_embeds = _build_comparison_embeds(session)
            await interaction.response.edit_message(content=None, embeds=info_embeds, view=self)
            await _update_paradigms(interaction.channel, session, para_embeds)


def _build_comparison_embeds(session: PrefSession) -> tuple[list[discord.Embed], list[discord.Embed]]:
    """Return (button_embeds, paradigm_embeds) for the current matchup.

    button_embeds: compact info card sent with the view/buttons.
    paradigm_embeds: one per judge, sent as separate messages.
    """
    pair = session.ranker.next_pair()
    if not pair:
        return [discord.Embed(title="✅ All comparisons done!", color=SUCCESS_COLOR)], []

    judge_a, judge_b = pair
    done = session.ranker.comparisons_done
    total = done + session.ranker.remaining
    anchors = session.ranker.anchors if session.ranker else []

    # --- Main info embed with side-by-side fields ---
    name_a = judge_a["name"]
    name_b = judge_b["name"]
    tag_a = f" *(known: {judge_a['score']})*" if judge_a.get("score") and judge_a in anchors else ""
    tag_b = f" *(known: {judge_b['score']})*" if judge_b.get("score") and judge_b in anchors else ""
    url_a = tabroom_paradigm_url(name_a)
    url_b = tabroom_paradigm_url(name_b)

    info = discord.Embed(title=f"⚔️ Matchup {done + 1}/{total}", color=EMBED_COLOR)
    info.add_field(name="🅰️ Judge A", value=f"**{name_a}**{tag_a}", inline=True)
    info.add_field(name="🅱️ Judge B", value=f"**{name_b}**{tag_b}", inline=True)
    info.add_field(name="\u200b", value="\u200b", inline=True)

    info.add_field(name="🏫 School", value=judge_a.get("school") or "?", inline=True)
    info.add_field(name="🏫 School", value=judge_b.get("school") or "?", inline=True)
    info.add_field(name="\u200b", value="\u200b", inline=True)

    info.add_field(name="🔄 Rounds", value=str(judge_a.get("rounds", "?")), inline=True)
    info.add_field(name="🔄 Rounds", value=str(judge_b.get("rounds", "?")), inline=True)
    info.add_field(name="\u200b", value="\u200b", inline=True)

    info.add_field(name="🔗 Tabroom", value=f"[View]({url_a})", inline=True)
    info.add_field(name="🔗 Tabroom", value=f"[View]({url_b})", inline=True)
    info.add_field(name="\u200b", value="\u200b", inline=True)

    # Count strikes/conflicts with round totals (pairwise + preloaded Notion data)
    all_session_judges = {j["name"]: j for j in session.unmatched}
    for csv_j, _, _ in session.matched:
        all_session_judges[csv_j["name"]] = csv_j

    # Start with pairwise-assigned strikes/conflicts
    strike_names = set(n for n, v in session.scores_map.items() if v == 6.0)
    conflict_names = set(n for n, v in session.scores_map.items() if v == 7.0)
    # Also include matched judges whose Notion score is already strike/conflict
    for csv_j, _, notion_score in session.matched:
        if csv_j["name"] in session.scores_map:
            continue  # pairwise override takes precedence
        if notion_score is not None and notion_score >= 7.0:
            conflict_names.add(csv_j["name"])
        elif notion_score is not None and notion_score >= 6.0:
            strike_names.add(csv_j["name"])

    strike_rounds = sum(all_session_judges.get(n, {}).get("rounds", 0) for n in strike_names)
    conflict_rounds = sum(all_session_judges.get(n, {}).get("rounds", 0) for n in conflict_names)
    footer_parts = [f"Matchup {done + 1}/{total}"]
    if hasattr(session.ranker, '_current_round'):
        footer_parts.append(f"Swiss round {session.ranker._current_round}/{session.ranker._rounds_total}")
    if strike_names:
        footer_parts.append(f"🚫 {len(strike_names)} strike{'s' if len(strike_names) != 1 else ''} ({strike_rounds} rds)")
    if conflict_names:
        footer_parts.append(f"⚠️ {len(conflict_names)} conflict{'s' if len(conflict_names) != 1 else ''} ({conflict_rounds} rds)")
    footer_parts.append("Press Done to finish early")
    info.set_footer(text=" • ".join(footer_parts))

    # --- Paradigm embeds (sent as separate messages) ---
    paradigm_embeds = []
    for side, judge, color in [("🅰️", judge_a, 0x5865F2), ("🅱️", judge_b, 0x57F287)]:
        paradigm = session.paradigms.get(judge["name"])
        if paradigm is None and tabroom_scraper and tabroom_cache:
            paradigm = tabroom_cache.get_or_fetch(judge["name"], tabroom_scraper)
            session.paradigms[judge["name"]] = paradigm

        phil = ""
        if paradigm and paradigm.get("philosophy"):
            phil = paradigm["philosophy"].strip()

        url = tabroom_paradigm_url(judge["name"])
        if phil:
            suffix = f"\n\n*… [read full on Tabroom →]({url})*"
            max_len = 4096 - len(suffix)
            if len(phil) > max_len:
                phil = phil[:max_len] + suffix
            pe = discord.Embed(title=f"{side} {judge['name']} — Paradigm", description=phil, color=color)
        else:
            pe = discord.Embed(
                title=f"{side} {judge['name']} — Paradigm",
                description=f"*No paradigm found.* [View on Tabroom →]({url})",
                color=color)
        paradigm_embeds.append(pe)

    return [info], paradigm_embeds


def _build_paradigm_embed(judge: dict, session: PrefSession, color: int = EMBED_COLOR) -> discord.Embed:
    """Build a paradigm embed for a single judge (same style as pairwise)."""
    paradigm = session.paradigms.get(judge["name"])
    if paradigm is None and tabroom_scraper and tabroom_cache:
        paradigm = tabroom_cache.get_or_fetch(judge["name"], tabroom_scraper)
        session.paradigms[judge["name"]] = paradigm

    url = tabroom_paradigm_url(judge["name"])
    phil = ""
    if paradigm and paradigm.get("philosophy"):
        phil = paradigm["philosophy"].strip()
    if phil:
        suffix = f"\n\n*… [read full on Tabroom →]({url})*"
        max_len = 4096 - len(suffix)
        if len(phil) > max_len:
            phil = phil[:max_len] + suffix
        return discord.Embed(title=f"📖 {judge['name']} — Paradigm", description=phil, color=color)
    else:
        return discord.Embed(
            title=f"📖 {judge['name']} — Paradigm",
            description=f"*No paradigm found.* [View on Tabroom →]({url})",
            color=color)


# ---------------------------------------------------------------------------
# Ordinal Ranking Views
# ---------------------------------------------------------------------------

class BulkStrikeSelect(ui.Select):
    """Multi-select dropdown for bulk strike/conflict assignment."""
    def __init__(self, session: PrefSession, judges: list[dict], action: str, page: int = 0):
        self.session = session
        self.action = action  # "strike" or "conflict"
        self.page = page
        label = "Strike" if action == "strike" else "Conflict"
        emoji = "🚫" if action == "strike" else "⚠️"
        start = page * 25
        batch = judges[start:start + 25]
        options = [discord.SelectOption(label=j["name"][:100], value=j["name"][:100],
                                        description=j.get("school", "")[:100])
                   for j in batch]
        super().__init__(placeholder=f"Select judges to {label}… (page {page + 1})",
                         options=options, min_values=0, max_values=len(options))

    async def callback(self, interaction: discord.Interaction):
        score = 6.0 if self.action == "strike" else 7.0
        for name in self.values:
            self.session.scores_map[name] = score
        label = "struck" if self.action == "strike" else "conflicted"
        count = len(self.values)
        if count > 0:
            await interaction.response.send_message(
                f"✅ **{count}** judge(s) marked as {label}.", ephemeral=True)
        else:
            await interaction.response.send_message("No judges selected.", ephemeral=True)


class BulkStrikeView(ui.View):
    """View for bulk strike/conflict selection before ordinal bucketing."""
    def __init__(self, session: PrefSession):
        super().__init__(timeout=None)
        self.session = session
        # Get judges not yet struck/conflicted
        available = [j for j in session.unmatched
                     if session.scores_map.get(j["name"], 0) < 6.0]
        # Add paginated strike selects (max 25 options per select, max 4 selects)
        pages = (len(available) + 24) // 25
        for page in range(min(pages, 2)):
            self.add_item(BulkStrikeSelect(session, available, "strike", page))
        # Conflict gets its own select if space permits
        if pages <= 2:
            for page in range(min(pages, 1)):
                self.add_item(BulkStrikeSelect(session, available, "conflict", page))

    @ui.button(label="✅ Done — Start Bucketing", style=discord.ButtonStyle.success, row=4)
    async def done(self, interaction: discord.Interaction, button: ui.Button):
        self.stop()
        session = self.session
        # Remove struck/conflicted judges from the bucketing pool
        struck = [n for n, s in session.scores_map.items() if s >= 6.0]
        session.unmatched = [j for j in session.unmatched
                             if session.scores_map.get(j["name"], 0) < 6.0]
        count = len(struck)
        desc = f"🚫 **{count}** judge(s) struck/conflicted." if count else "No strikes or conflicts."
        desc += f"\n📊 **{len(session.unmatched)}** judges to bucket into tiers."
        await interaction.response.edit_message(
            embed=discord.Embed(description=desc, color=EMBED_COLOR), view=None)
        session.state = "ordinal_bucketing"
        session.current_idx = 0
        await _send_ordinal_bucket(interaction.channel, session)

    @ui.button(label="⏭ Skip — No Strikes", style=discord.ButtonStyle.secondary, row=4)
    async def skip(self, interaction: discord.Interaction, button: ui.Button):
        self.stop()
        session = self.session
        await interaction.response.edit_message(
            embed=discord.Embed(
                description=f"📊 **{len(session.unmatched)}** judges to bucket into tiers.",
                color=EMBED_COLOR), view=None)
        session.state = "ordinal_bucketing"
        session.current_idx = 0
        await _send_ordinal_bucket(interaction.channel, session)


class OrdinalBucketView(ui.View):
    """Fast one-tap tier assignment for each judge."""
    def __init__(self, session: PrefSession):
        super().__init__(timeout=None)
        self.session = session

    @ui.button(label="1 Best", style=discord.ButtonStyle.success, row=0)
    async def tier1(self, interaction: discord.Interaction, button: ui.Button):
        await self._assign(interaction, 1.0)

    @ui.button(label="2 Good", style=discord.ButtonStyle.success, row=0)
    async def tier2(self, interaction: discord.Interaction, button: ui.Button):
        await self._assign(interaction, 2.0)

    @ui.button(label="3 Mid", style=discord.ButtonStyle.primary, row=0)
    async def tier3(self, interaction: discord.Interaction, button: ui.Button):
        await self._assign(interaction, 3.0)

    @ui.button(label="4 Bad", style=discord.ButtonStyle.secondary, row=0)
    async def tier4(self, interaction: discord.Interaction, button: ui.Button):
        await self._assign(interaction, 4.0)

    @ui.button(label="5 Worst", style=discord.ButtonStyle.secondary, row=0)
    async def tier5(self, interaction: discord.Interaction, button: ui.Button):
        await self._assign(interaction, 5.0)

    @ui.button(label="Strike", style=discord.ButtonStyle.danger, row=1)
    async def strike(self, interaction: discord.Interaction, button: ui.Button):
        await self._assign(interaction, 6.0)

    @ui.button(label="Conflict", style=discord.ButtonStyle.secondary, row=1)
    async def conflict(self, interaction: discord.Interaction, button: ui.Button):
        await self._assign(interaction, 7.0)

    @ui.button(label="◀ Previous", style=discord.ButtonStyle.secondary, row=1)
    async def prev(self, interaction: discord.Interaction, button: ui.Button):
        if self.session.current_idx > 0:
            self.session.current_idx -= 1
        await _show_ordinal_bucket_judge(interaction, self.session, self)

    @ui.button(label="💾 Save", style=discord.ButtonStyle.secondary, row=1)
    async def save_btn(self, interaction: discord.Interaction, button: ui.Button):
        file_bytes, filename = save_progress(self.session)
        file = discord.File(io.BytesIO(file_bytes), filename=filename)
        await interaction.response.send_message(
            embed=discord.Embed(title="💾 Progress Saved",
                                description="Download to resume later.",
                                color=SUCCESS_COLOR),
            file=file, ephemeral=True)

    @ui.button(label="🔍 Compare", style=discord.ButtonStyle.primary, row=1)
    async def compare_btn(self, interaction: discord.Interaction, button: ui.Button):
        current_judge = self.session.unmatched[self.session.current_idx]
        all_names = [j["name"] for j in self.session.csv_judges if j["name"] != current_judge["name"]]
        options = [discord.SelectOption(label=n[:100], value=n[:100]) for n in all_names[:25]]
        if not options:
            await interaction.response.send_message("No other judges to compare.", ephemeral=True)
            return
        view = OrdinalCompareSelectView(self.session, current_judge, options, parent_view=self)
        await interaction.response.send_message("Pick a judge to compare with:", view=view, ephemeral=True)

    async def _assign(self, interaction: discord.Interaction, score: float):
        session = self.session
        judge = session.unmatched[session.current_idx]
        session.scores_map[judge["name"]] = score
        session.current_idx += 1
        if session.current_idx < len(session.unmatched):
            await _show_ordinal_bucket_judge(interaction, session, self)
        else:
            self.stop()
            await _finish_ordinal_bucketing(interaction, session)


class OrdinalCompareSelectView(ui.View):
    """Select a judge to compare side-by-side during ordinal bucketing."""
    def __init__(self, session: PrefSession, current_judge: dict,
                 options: list[discord.SelectOption], parent_view: OrdinalBucketView):
        super().__init__(timeout=None)
        self.session = session
        self.current_judge = current_judge
        self.parent_view = parent_view
        select = ui.Select(placeholder="Select a judge to compare…", options=options)
        select.callback = self.on_select
        self.add_item(select)

    async def on_select(self, interaction: discord.Interaction):
        selected_name = interaction.data["values"][0]
        other = next((j for j in self.session.csv_judges if j["name"] == selected_name), None)
        if not other:
            await interaction.response.send_message("Judge not found.", ephemeral=True)
            return
        score1 = self.session.scores_map.get(self.current_judge["name"])
        score2 = self.session.scores_map.get(other["name"])
        p1 = self.session.paradigms.get(self.current_judge["name"])
        p2 = self.session.paradigms.get(other["name"])
        embed1 = build_judge_embed(self.current_judge, p1, score=score1)
        embed1.title = f"🅰️ {self.current_judge['name']}"
        embed2 = build_judge_embed(other, p2, score=score2)
        embed2.title = f"🅱️ {other['name']}"
        await interaction.response.edit_message(
            content="**Side-by-side comparison:**", embeds=[embed1, embed2], view=None)


class OrdinalRefinePromptView(ui.View):
    """Show tier summary and let user pick which tier to refine via pairwise."""
    def __init__(self, session: PrefSession):
        super().__init__(timeout=None)
        self.session = session

    @ui.button(label="Refine Tier 1", style=discord.ButtonStyle.success, row=0)
    async def refine_1(self, interaction: discord.Interaction, button: ui.Button):
        await self._refine(interaction, 1)

    @ui.button(label="Refine Tier 2", style=discord.ButtonStyle.success, row=0)
    async def refine_2(self, interaction: discord.Interaction, button: ui.Button):
        await self._refine(interaction, 2)

    @ui.button(label="Refine Tier 3", style=discord.ButtonStyle.primary, row=0)
    async def refine_3(self, interaction: discord.Interaction, button: ui.Button):
        await self._refine(interaction, 3)

    @ui.button(label="Refine Tier 4", style=discord.ButtonStyle.secondary, row=0)
    async def refine_4(self, interaction: discord.Interaction, button: ui.Button):
        await self._refine(interaction, 4)

    @ui.button(label="Refine Tier 5", style=discord.ButtonStyle.secondary, row=0)
    async def refine_5(self, interaction: discord.Interaction, button: ui.Button):
        await self._refine(interaction, 5)

    @ui.button(label="✅ Finish — Export Rankings", style=discord.ButtonStyle.success, row=1)
    async def finish(self, interaction: discord.Interaction, button: ui.Button):
        self.stop()
        await interaction.response.edit_message(
            embed=discord.Embed(description="⏳ Building export…", color=EMBED_COLOR), view=None)
        await _export_ordinal_rankings(interaction.channel, self.session)

    @ui.button(label="💾 Save Progress", style=discord.ButtonStyle.secondary, row=1)
    async def save_btn(self, interaction: discord.Interaction, button: ui.Button):
        file_bytes, filename = save_progress(self.session)
        file = discord.File(io.BytesIO(file_bytes), filename=filename)
        await interaction.response.send_message(
            embed=discord.Embed(title="💾 Progress Saved",
                                description="Download to resume later.",
                                color=SUCCESS_COLOR),
            file=file, ephemeral=True)

    async def _refine(self, interaction: discord.Interaction, tier: int):
        session = self.session
        tier_judges = [j for j in session.csv_judges
                       if session.scores_map.get(j["name"]) == float(tier)]
        if len(tier_judges) < 2:
            await interaction.response.send_message(
                f"Tier {tier} has {len(tier_judges)} judge(s) — nothing to refine.", ephemeral=True)
            return
        # Check if already refined
        if tier in session.ordinal_rankings and session.ordinal_rankings[tier]:
            await interaction.response.send_message(
                f"Tier {tier} already refined ({len(session.ordinal_rankings[tier])} judges). "
                "Press **Finish** to export or refine a different tier.", ephemeral=True)
            return
        self.stop()
        session.ordinal_refine_tier = tier
        session.state = "ordinal_refining"
        # Build pairwise ranker for just this tier
        session.ranker = PairwiseRanker(tier_judges, anchor_judges=[])
        await interaction.response.edit_message(
            embed=discord.Embed(
                description=f"🔀 Refining **Tier {tier}** — "
                            f"**{session.ranker.total_comparisons}** matchups for "
                            f"**{len(tier_judges)}** judges.",
                color=EMBED_COLOR), view=None)
        info_embeds, para_embeds = _build_comparison_embeds(session)
        view = OrdinalPairwiseView(session)
        await interaction.channel.send(embeds=info_embeds, view=view)
        await _update_paradigms(interaction.channel, session, para_embeds)


class OrdinalPairwiseView(ui.View):
    """Pairwise comparison view for within-tier ordinal refinement."""
    def __init__(self, session: PrefSession):
        super().__init__(timeout=None)
        self.session = session

    @ui.button(label="A", style=discord.ButtonStyle.primary, row=0)
    async def pick_a(self, interaction: discord.Interaction, button: ui.Button):
        await self._record(interaction, "a")

    @ui.button(label="B", style=discord.ButtonStyle.success, row=0)
    async def pick_b(self, interaction: discord.Interaction, button: ui.Button):
        await self._record(interaction, "b")

    @ui.button(label="Skip", style=discord.ButtonStyle.secondary, row=0)
    async def skip(self, interaction: discord.Interaction, button: ui.Button):
        await self._record(interaction, "skip")

    @ui.button(label="Undo", style=discord.ButtonStyle.secondary, row=0)
    async def undo(self, interaction: discord.Interaction, button: ui.Button):
        session = self.session
        if not session.ranker:
            return
        if session.ranker.undo():
            info_embeds, para_embeds = _build_comparison_embeds(session)
            await interaction.response.edit_message(content="↩️ Undone.", embeds=info_embeds, view=self)
            await _update_paradigms(interaction.channel, session, para_embeds)
        else:
            await interaction.response.edit_message(content="Nothing to undo.")

    @ui.button(label="Done", style=discord.ButtonStyle.secondary, row=0)
    async def done(self, interaction: discord.Interaction, button: ui.Button):
        self.stop()
        await interaction.response.edit_message(content="⏳ Finishing tier refinement…",
                                                embeds=[], view=None)
        await _finish_ordinal_tier_refine(interaction.channel, self.session)

    @ui.button(label="💾 Save", style=discord.ButtonStyle.secondary, row=1)
    async def save_btn(self, interaction: discord.Interaction, button: ui.Button):
        file_bytes, filename = save_progress(self.session)
        file = discord.File(io.BytesIO(file_bytes), filename=filename)
        await interaction.response.send_message(
            embed=discord.Embed(title="💾 Progress Saved",
                                description="Download to resume later.",
                                color=SUCCESS_COLOR),
            file=file, ephemeral=True)

    async def _record(self, interaction: discord.Interaction, choice: str):
        session = self.session
        if not session.ranker:
            return
        pair = session.ranker.next_pair()
        if not pair:
            self.stop()
            await interaction.response.edit_message(content="⏳ Finishing tier refinement…",
                                                    embeds=[], view=None)
            await _finish_ordinal_tier_refine(interaction.channel, session)
            return
        judge_a, judge_b = pair
        if choice == "a":
            session.ranker.record_result(judge_a, judge_b)
        elif choice == "b":
            session.ranker.record_result(judge_b, judge_a)
        else:
            session.ranker.skip_pair()
        if session.ranker.is_complete:
            self.stop()
            await interaction.response.edit_message(content="⏳ Finishing tier refinement…",
                                                    embeds=[], view=None)
            await _finish_ordinal_tier_refine(interaction.channel, session)
        else:
            info_embeds, para_embeds = _build_comparison_embeds(session)
            await interaction.response.edit_message(content=None, embeds=info_embeds, view=self)
            await _update_paradigms(interaction.channel, session, para_embeds)


# --- Ordinal helper functions ---

async def _send_ordinal_bulk_strike(channel, session: PrefSession):
    """Send the bulk strike/conflict selection view."""
    total = len(session.unmatched)
    embed = discord.Embed(
        title="🚫 Bulk Strike / Conflict",
        description=f"**{total}** judges in the pool.\n"
                    "Select any judges you want to **strike** or **conflict** in bulk.\n"
                    "Then press **Done** to start tier bucketing.\n\n"
                    "*You can also skip this step if you have no strikes.*",
        color=EMBED_COLOR)
    view = BulkStrikeView(session)
    await channel.send(embed=embed, view=view)


async def _send_ordinal_bucket(channel, session: PrefSession):
    """Send the bucketing view for the current judge."""
    if session.current_idx >= len(session.unmatched):
        await _finish_ordinal_bucketing_from_channel(channel, session)
        return
    judge = session.unmatched[session.current_idx]
    idx = session.current_idx + 1
    total = len(session.unmatched)
    existing = session.scores_map.get(judge["name"])

    embed = discord.Embed(title=f"Judge {idx}/{total} — {judge['name']}", color=EMBED_COLOR)
    embed.add_field(name="🏫 School", value=judge.get("school") or "Unknown", inline=True)
    embed.add_field(name="🔄 Rounds", value=str(judge.get("rounds", "?")), inline=True)
    if existing is not None and existing < 6.0:
        embed.add_field(name="📋 Current Tier", value=str(int(existing)), inline=True)
    url = tabroom_paradigm_url(judge["name"])
    embed.add_field(name="🔗 Tabroom", value=f"[View full paradigm]({url})", inline=False)

    # Show tier distribution so far
    tier_counts = {}
    for name, score in session.scores_map.items():
        if score < 6.0:
            t = int(score)
            tier_counts[t] = tier_counts.get(t, 0) + 1
    dist_parts = [f"T{t}: {c}" for t, c in sorted(tier_counts.items())]
    strikes = sum(1 for s in session.scores_map.values() if s == 6.0)
    conflicts = sum(1 for s in session.scores_map.values() if s == 7.0)
    footer = f"Progress: {idx}/{total}"
    if dist_parts:
        footer += " • " + " | ".join(dist_parts)
    if strikes:
        footer += f" | 🚫{strikes}"
    if conflicts:
        footer += f" | ⚠️{conflicts}"
    embed.set_footer(text=footer)

    view = OrdinalBucketView(session)
    await channel.send(embed=embed, view=view)
    para_embed = _build_paradigm_embed(judge, session)
    await _update_paradigms(channel, session, [para_embed])


async def _show_ordinal_bucket_judge(interaction: discord.Interaction,
                                     session: PrefSession, view: OrdinalBucketView):
    """Update the embed for the current judge during bucketing."""
    judge = session.unmatched[session.current_idx]
    idx = session.current_idx + 1
    total = len(session.unmatched)
    existing = session.scores_map.get(judge["name"])

    embed = discord.Embed(title=f"Judge {idx}/{total} — {judge['name']}", color=EMBED_COLOR)
    embed.add_field(name="🏫 School", value=judge.get("school") or "Unknown", inline=True)
    embed.add_field(name="🔄 Rounds", value=str(judge.get("rounds", "?")), inline=True)
    if existing is not None and existing < 6.0:
        embed.add_field(name="📋 Current Tier", value=str(int(existing)), inline=True)
    url = tabroom_paradigm_url(judge["name"])
    embed.add_field(name="🔗 Tabroom", value=f"[View full paradigm]({url})", inline=False)

    tier_counts = {}
    for name, score in session.scores_map.items():
        if score < 6.0:
            t = int(score)
            tier_counts[t] = tier_counts.get(t, 0) + 1
    dist_parts = [f"T{t}: {c}" for t, c in sorted(tier_counts.items())]
    strikes = sum(1 for s in session.scores_map.values() if s == 6.0)
    conflicts = sum(1 for s in session.scores_map.values() if s == 7.0)
    footer = f"Progress: {idx}/{total}"
    if dist_parts:
        footer += " • " + " | ".join(dist_parts)
    if strikes:
        footer += f" | 🚫{strikes}"
    if conflicts:
        footer += f" | ⚠️{conflicts}"
    embed.set_footer(text=footer)

    await interaction.response.edit_message(embed=embed, view=view)
    para_embed = _build_paradigm_embed(judge, session)
    await _update_paradigms(interaction.channel, session, [para_embed])


async def _finish_ordinal_bucketing(interaction: discord.Interaction, session: PrefSession):
    """Transition from bucketing to the refine prompt."""
    # Clean up paradigm messages
    for msg in session.paradigm_messages:
        try:
            await msg.delete()
        except Exception:
            pass
    session.paradigm_messages.clear()

    session.state = "ordinal_refine_prompt"
    embed = _build_ordinal_summary_embed(session)
    view = OrdinalRefinePromptView(session)
    await interaction.response.edit_message(embed=embed, view=view)


async def _finish_ordinal_bucketing_from_channel(channel, session: PrefSession):
    """Same as above but sends to channel (for resume path)."""
    for msg in session.paradigm_messages:
        try:
            await msg.delete()
        except Exception:
            pass
    session.paradigm_messages.clear()

    session.state = "ordinal_refine_prompt"
    embed = _build_ordinal_summary_embed(session)
    view = OrdinalRefinePromptView(session)
    await channel.send(embed=embed, view=view)


def _build_ordinal_summary_embed(session: PrefSession) -> discord.Embed:
    """Build the tier summary embed for the refine prompt."""
    embed = discord.Embed(
        title="📊 Tier Summary — Choose Tiers to Refine",
        description="Select a tier to rank judges within it via pairwise comparison.\n"
                    "Press **Finish** when you're done to export the final ordinal rankings.",
        color=SUCCESS_COLOR)

    for tier in range(1, 6):
        names = [j["name"] for j in session.csv_judges
                 if session.scores_map.get(j["name"]) == float(tier)]
        refined = tier in session.ordinal_rankings and session.ordinal_rankings[tier]
        status = "✅ Refined" if refined else "⬜ Not refined"
        value = f"**{len(names)}** judges — {status}"
        if refined:
            # Show first few ranked names
            ranked = session.ordinal_rankings[tier]
            preview = ", ".join(ranked[:5])
            if len(ranked) > 5:
                preview += f"… (+{len(ranked) - 5})"
            value += f"\n{preview}"
        embed.add_field(name=f"Tier {tier}", value=value, inline=False)

    strikes = sum(1 for s in session.scores_map.values() if s == 6.0)
    conflicts = sum(1 for s in session.scores_map.values() if s == 7.0)
    if strikes or conflicts:
        embed.add_field(name="🚫 Strikes / ⚠️ Conflicts",
                        value=f"{strikes} strike(s), {conflicts} conflict(s)", inline=False)
    return embed


async def _finish_ordinal_tier_refine(channel, session: PrefSession):
    """Save pairwise results for the current tier and return to refine prompt."""
    tier = session.ordinal_refine_tier
    if session.ranker:
        rankings = session.ranker.get_rankings()
        session.ordinal_rankings[tier] = [j["name"] for j, elo in rankings]

        embed = discord.Embed(
            title=f"✅ Tier {tier} Refined",
            description=f"**{len(rankings)}** judges ranked within Tier {tier}.",
            color=SUCCESS_COLOR)
        lines = []
        for rank, (j, elo) in enumerate(rankings, 1):
            lines.append(f"`#{rank}` **{j['name']}** (Elo: {elo:.0f})")
        text = "\n".join(lines[:20])
        if len(lines) > 20:
            text += f"\n… and {len(lines) - 20} more"
        embed.add_field(name="Rankings", value=text or "None", inline=False)
        await channel.send(embed=embed)

    session.ranker = None
    session.ordinal_refine_tier = None

    # Clean up paradigm messages
    for msg in session.paradigm_messages:
        try:
            await msg.delete()
        except Exception:
            pass
    session.paradigm_messages.clear()

    session.state = "ordinal_refine_prompt"
    summary = _build_ordinal_summary_embed(session)
    view = OrdinalRefinePromptView(session)
    await channel.send(embed=summary, view=view)


async def _export_ordinal_rankings(channel, session: PrefSession):
    """Export final ordinal rankings as a formatted Excel file."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ordinal Rankings"

    # Styling
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="5865F2", end_color="5865F2", fill_type="solid")
    tier_colors = {
        1: PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),  # green
        2: PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid"),  # light blue
        3: PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),  # yellow
        4: PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),  # orange
        5: PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid"),  # red-ish
        6: PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),  # grey
        7: PatternFill(start_color="BFBFBF", end_color="BFBFBF", fill_type="solid"),  # dark grey
    }
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    headers = ["Rank", "Name", "School", "Rounds", "Tier"]
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 10

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    judge_lookup = {j["name"]: j for j in session.csv_judges}
    row = 2
    global_rank = 1

    for tier in range(1, 6):
        tier_names = [j["name"] for j in session.csv_judges
                      if session.scores_map.get(j["name"]) == float(tier)]
        if not tier_names:
            continue

        # Use refined order if available, otherwise alphabetical
        if tier in session.ordinal_rankings and session.ordinal_rankings[tier]:
            ordered = session.ordinal_rankings[tier]
            # Add any tier members not in the refined list (edge case)
            refined_set = set(ordered)
            for name in tier_names:
                if name not in refined_set:
                    ordered.append(name)
        else:
            ordered = sorted(tier_names)

        for name in ordered:
            j = judge_lookup.get(name, {})
            fill = tier_colors.get(tier, PatternFill())
            for col, val in enumerate([global_rank, name, j.get("school", ""),
                                        j.get("rounds", 0), f"Tier {tier}"], 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.fill = fill
                cell.border = thin_border
                if col == 1:
                    cell.alignment = Alignment(horizontal="center")
            row += 1
            global_rank += 1

    # Strikes
    strike_names = sorted(n for n, s in session.scores_map.items() if s == 6.0)
    for name in strike_names:
        j = judge_lookup.get(name, {})
        fill = tier_colors[6]
        for col, val in enumerate(["S", name, j.get("school", ""),
                                    j.get("rounds", 0), "Strike"], 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill = fill
            cell.border = thin_border
            if col == 1:
                cell.alignment = Alignment(horizontal="center")
        row += 1

    # Conflicts
    conflict_names = sorted(n for n, s in session.scores_map.items() if s == 7.0)
    for name in conflict_names:
        j = judge_lookup.get(name, {})
        fill = tier_colors[7]
        for col, val in enumerate(["C", name, j.get("school", ""),
                                    j.get("rounds", 0), "Conflict"], 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill = fill
            cell.border = thin_border
            if col == 1:
                cell.alignment = Alignment(horizontal="center")
        row += 1

    # Write to bytes
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    file = discord.File(buf, filename="ordinal_rankings.xlsx")

    total_ranked = global_rank - 1
    embed = discord.Embed(
        title="✅ Ordinal Rankings Complete!",
        description=f"**{total_ranked}** judges ranked, "
                    f"**{len(strike_names)}** strikes, "
                    f"**{len(conflict_names)}** conflicts.\n"
                    "Here's your ordinal ranking sheet.",
        color=SUCCESS_COLOR)
    await channel.send(embed=embed, file=file)
    session.state = "done"
    sessions.pop(channel.id, None)


# ---------------------------------------------------------------------------
# Quota Views
# ---------------------------------------------------------------------------

class QuotaModeView(ui.View):
    def __init__(self, session: PrefSession):
        super().__init__(timeout=None)
        self.session = session

    @ui.button(label="1️⃣ Round Count", style=discord.ButtonStyle.primary)
    async def rounds(self, interaction: discord.Interaction, button: ui.Button):
        self.session.quota_mode = "rounds"
        self.stop()
        await self._ask_quotas(interaction)

    @ui.button(label="2️⃣ Judge Count", style=discord.ButtonStyle.primary)
    async def judges(self, interaction: discord.Interaction, button: ui.Button):
        self.session.quota_mode = "judges"
        self.stop()
        await self._ask_quotas(interaction)

    async def _ask_quotas(self, interaction: discord.Interaction):
        self.session.state = "awaiting_quotas"
        unit = "judges" if self.session.quota_mode == "judges" else "rounds"
        embed = discord.Embed(
            title="📝 Tier Quotas",
            description=f"Enter **minimum {unit}** for each tier in one line:\n"
                        f"```\nTier1  Tier2  Tier3  Tier4  Tier5  Strike\n```\n"
                        f"Use `-` to skip a tier. Use `min,max` for a range.\n\n"
                        f"**Example:** `6 8 10 8 6 -`",
            color=EMBED_COLOR)
        embed.set_footer(text="Strike tier is optional • e.g. '6 8 10 8 6' also works")
        await interaction.response.edit_message(embed=embed, view=None)


# ---------------------------------------------------------------------------
# Events
# ---------------------------------------------------------------------------

@client.event
async def on_ready():
    global tabroom_scraper
    print(f"Prefessor Judge is online as {client.user}", flush=True)
    if TabroomScraper:
        tabroom_scraper = TabroomScraper()
        tabroom_scraper.login()


@client.event
async def on_message(message: discord.Message):
    if message.author == client.user:
        return

    channel_id = message.channel.id
    session = sessions.get(channel_id)

    if client.user.mentioned_in(message) and not session:
        content = message.content.lower()
        if "cancel" in content:
            # No active session to cancel
            await message.channel.send(embed=discord.Embed(
                description="No active prefs session to cancel.", color=WARN_COLOR))
            return
        if any(kw in content for kw in ("pref", "judge", "rank", "tournament", "hello", "hi")):
            session = PrefSession(channel_id, message.author.id)
            sessions[channel_id] = session
            # Check for a progress file (.xlsx) to resume from
            xlsx_attachment = None
            csv_attachment = None
            for att in message.attachments:
                if att.filename.endswith(".xlsx"):
                    xlsx_attachment = att
                elif att.filename.endswith(".csv"):
                    csv_attachment = att
            if xlsx_attachment:
                xlsx_bytes = await xlsx_attachment.read()
                if is_progress_file(xlsx_bytes):
                    await _resume_from_file(message, session, xlsx_attachment, file_bytes=xlsx_bytes)
                    return
                # Not a progress file — treat as regular attachment, fall through to CSV check
                xlsx_attachment = None
            # If CSV attached with the command, process it immediately
            if csv_attachment:
                csv_bytes = await csv_attachment.read()
                tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".csv", mode="wb")
                tmp.write(csv_bytes)
                tmp.close()
                try:
                    session.csv_judges = parse_tournament_csv(tmp.name)
                finally:
                    os.unlink(tmp.name)
                total = len(session.csv_judges)
                await message.channel.send(embed=discord.Embed(
                    description=f"✅ Loaded **{total}** judges from `{csv_attachment.filename}`.",
                    color=EMBED_COLOR))
                session.state = "awaiting_rating_range"
                await _send_rating_range_prompt(message.channel, session)
            else:
                embed = discord.Embed(
                    title="📋 Prefessor Judge",
                    description="Let's do prefs! Upload the tournament judge CSV file.\n\n"
                                "Expected columns: `Name, School, Rounds, Your Rating`\n"
                                "(or `First, Last, School, Online, Rounds, Rating`)\n\n"
                                "💾 **Resuming?** Upload a `.xlsx` progress file instead.",
                    color=EMBED_COLOR)
                embed.set_footer(text="Type 'cancel' at any time to abort.")
                await message.channel.send(embed=embed)
            return
        else:
            await message.channel.send(embed=discord.Embed(
                title="👋 Prefessor Judge",
                description='Mention me and say something like **"do prefs"** to start!',
                color=EMBED_COLOR))
            return

    if not session:
        return
    if message.author.id != session.user_id:
        return

    # Strip bot mention from content for command parsing
    raw = message.content.strip().lower()
    clean = raw.replace(f"<@{client.user.id}>", "").replace(f"<@!{client.user.id}>", "").strip()
    if clean == "cancel" or raw == "cancel":
        del sessions[channel_id]
        await message.channel.send(embed=discord.Embed(
            title="❌ Cancelled", description="Prefs workflow cancelled.", color=ERROR_COLOR))
        return

    if clean == "resume" or raw == "resume":
        if session.state == "comparing" and session.ranker and not session.ranker.is_complete:
            info_embeds, para_embeds = _build_comparison_embeds(session)
            view = PairwiseView(session)
            await _update_paradigms(message.channel, session, para_embeds)
            await message.channel.send(
                content=f"▶️ Resuming — {session.ranker.remaining} comparisons left.",
                embeds=info_embeds, view=view)
        elif session.state == "prompting_scores":
            judge = session.unmatched[session.current_idx]
            total = len(session.unmatched)
            idx = session.current_idx + 1
            embed = discord.Embed(title=f"Judge {idx}/{total} — {judge['name']}", color=EMBED_COLOR)
            embed.add_field(name="🏫 School", value=judge.get("school") or "Unknown", inline=True)
            embed.add_field(name="🔄 Rounds", value=str(judge.get("rounds", "?")), inline=True)
            embed.set_footer(text=f"Progress: {idx}/{total}")
            view = ScoringView(session)
            await message.channel.send(content="▶️ Resuming scoring…", embed=embed, view=view)
            para_embed = _build_paradigm_embed(judge, session)
            await _update_paradigms(message.channel, session, [para_embed])
        elif session.state == "ordinal_bucketing" and session.current_idx < len(session.unmatched):
            await message.channel.send(content="▶️ Resuming tier bucketing…")
            await _send_ordinal_bucket(message.channel, session)
        elif session.state == "ordinal_refining" and session.ranker and not session.ranker.is_complete:
            tier = session.ordinal_refine_tier or "?"
            info_embeds, para_embeds = _build_comparison_embeds(session)
            view = OrdinalPairwiseView(session)
            await _update_paradigms(message.channel, session, para_embeds)
            await message.channel.send(
                content=f"▶️ Resuming Tier {tier} refinement — "
                        f"{session.ranker.remaining} comparisons left.",
                embeds=info_embeds, view=view)
        elif session.state == "ordinal_refine_prompt":
            embed = _build_ordinal_summary_embed(session)
            view = OrdinalRefinePromptView(session)
            await message.channel.send(embed=embed, view=view)
        else:
            await message.channel.send(embed=discord.Embed(
                description=f"Current state: **{session.state}** — nothing to resume. Send your input to continue.",
                color=WARN_COLOR))
        return

    try:
        await handle_state(message, session)
    except Exception as e:
        await message.channel.send(f"⚠️ Error: {e}")
        import traceback; traceback.print_exc()
        sessions.pop(channel_id, None)


# ---------------------------------------------------------------------------
# State Machine
# ---------------------------------------------------------------------------

async def handle_state(message: discord.Message, session: PrefSession):
    channel = message.channel

    # --- Awaiting CSV ---
    if session.state == "awaiting_csv":
        if not message.attachments:
            await channel.send(embed=discord.Embed(
                description="Please upload a CSV file (or an `.xlsx` progress file to resume).",
                color=WARN_COLOR))
            return
        attachment = message.attachments[0]
        if attachment.filename.endswith(".xlsx"):
            xlsx_bytes = await attachment.read()
            if is_progress_file(xlsx_bytes):
                await _resume_from_file(message, session, attachment, file_bytes=xlsx_bytes)
                return
            # Not a progress file — tell user to upload CSV
            await channel.send(embed=discord.Embed(
                description="That `.xlsx` file doesn't look like a saved progress file.\n"
                            "Please upload a `.csv` tournament file or a valid progress file.",
                color=ERROR_COLOR))
            return
        if not attachment.filename.endswith(".csv"):
            await channel.send(embed=discord.Embed(
                description="That doesn't look like a CSV or progress file. "
                            "Please upload a `.csv` or `.xlsx` file.", color=ERROR_COLOR))
            return
        csv_bytes = await attachment.read()
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".csv", mode="wb")
        tmp.write(csv_bytes)
        tmp.close()
        try:
            session.csv_judges = parse_tournament_csv(tmp.name)
        finally:
            os.unlink(tmp.name)

        total = len(session.csv_judges)
        await channel.send(embed=discord.Embed(
            description=f"✅ Loaded **{total}** judges from `{attachment.filename}`.",
            color=EMBED_COLOR))
        session.state = "awaiting_rating_range"
        await _send_rating_range_prompt(channel, session)

    # --- Awaiting rating range ---
    elif session.state == "awaiting_rating_range":
        text = message.content.strip()
        try:
            val = int(text)
            if val < 2 or val > 20:
                await channel.send(embed=discord.Embed(
                    description="Please enter a number between 2 and 20, or use the buttons.",
                    color=WARN_COLOR))
                return
            await _apply_rating_range(channel, session, val)
        except ValueError:
            await channel.send(embed=discord.Embed(
                description="Please enter a number (e.g. **5** for a 1-5 scale) or use the buttons.",
                color=WARN_COLOR))

    # --- Awaiting source choice (handled by buttons, text fallback) ---
    elif session.state == "awaiting_source_choice":
        await channel.send(embed=discord.Embed(
            description="Please use the buttons above to choose a rating source.", color=WARN_COLOR))

    # --- Awaiting unmatched choice (handled by buttons, text fallback) ---
    elif session.state == "awaiting_unmatched_choice":
        await channel.send(embed=discord.Embed(
            description="Please use the buttons above to choose how to handle unmatched judges.",
            color=WARN_COLOR))

    # --- Pairwise comparing (text fallback) ---
    elif session.state == "comparing":
        text = message.content.strip().lower()
        if not session.ranker:
            return
        pair = session.ranker.next_pair()
        if not pair:
            await _finish_comparison(channel, session)
            return
        judge_a, judge_b = pair
        if text == "a":
            session.ranker.record_result(judge_a, judge_b)
        elif text == "b":
            session.ranker.record_result(judge_b, judge_a)
        elif text in ("skip", "s"):
            session.ranker.skip_pair()
        elif text in ("done", "stop"):
            await _finish_comparison(channel, session)
            return
        else:
            await channel.send("Type **A**, **B**, **skip**, or **done**:")
            return
        if session.ranker.is_complete:
            await _finish_comparison(channel, session)
        else:
            info_embeds, para_embeds = _build_comparison_embeds(session)
            view = PairwiseView(session)
            await channel.send(embeds=info_embeds, view=view)
            await _update_paradigms(channel, session, para_embeds)

    # --- Awaiting quota text input ---
    elif session.state == "awaiting_quotas":
        text = message.content.strip()
        try:
            quotas = parse_all_quotas(text, session.quota_mode)
            if not quotas:
                await channel.send(embed=discord.Embed(
                    description="Invalid format. Enter one value per tier separated by spaces.\n"
                                "Example: `6 8 10 8 6 -`", color=ERROR_COLOR))
                return
            session.quotas = quotas
            # Show confirmation
            unit = "judges" if session.quota_mode == "judges" else "rounds"
            lines = []
            for t in range(1, 7):
                label = "S" if t == 6 else str(t)
                q = quotas.get(t)
                if q:
                    mn = q.get("min", "-")
                    mx = q.get("max", "-")
                    lines.append(f"Tier {label}: min={mn}, max={mx}")
                else:
                    lines.append(f"Tier {label}: —")
            await channel.send(embed=discord.Embed(
                title="✅ Quotas Set",
                description=f"**Mode:** {unit}\n" + "\n".join(lines),
                color=EMBED_COLOR))
            await run_assignment(channel, session)
        except ValueError:
            await channel.send(embed=discord.Embed(
                description="Invalid numbers. Use: `6 8 10 8 6 -`", color=ERROR_COLOR))

    # --- Done state ---
    elif session.state == "done":
        if client.user.mentioned_in(message):
            sessions.pop(channel.id, None)
            sessions[channel.id] = PrefSession(channel.id, message.author.id)
            await channel.send(embed=discord.Embed(
                title="📋 Prefessor Judge — New Session",
                description="Upload a tournament CSV to begin.", color=EMBED_COLOR))

    # --- Ordinal states (button-driven, text fallback) ---
    elif session.state in ("ordinal_bulk_strike", "ordinal_bucketing",
                           "ordinal_refine_prompt", "ordinal_refining"):
        text = message.content.strip().lower()
        if session.state == "ordinal_refining" and session.ranker:
            pair = session.ranker.next_pair()
            if not pair:
                await _finish_ordinal_tier_refine(channel, session)
                return
            judge_a, judge_b = pair
            if text == "a":
                session.ranker.record_result(judge_a, judge_b)
            elif text == "b":
                session.ranker.record_result(judge_b, judge_a)
            elif text in ("skip", "s"):
                session.ranker.skip_pair()
            elif text in ("done", "stop"):
                await _finish_ordinal_tier_refine(channel, session)
                return
            else:
                await channel.send("Type **A**, **B**, **skip**, or **done**:")
                return
            if session.ranker.is_complete:
                await _finish_ordinal_tier_refine(channel, session)
            else:
                info_embeds, para_embeds = _build_comparison_embeds(session)
                view = OrdinalPairwiseView(session)
                await channel.send(embeds=info_embeds, view=view)
                await _update_paradigms(channel, session, para_embeds)
        else:
            await channel.send(embed=discord.Embed(
                description="Please use the buttons above to continue.",
                color=WARN_COLOR))


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

async def _update_paradigms(channel, session: PrefSession, para_embeds: list):
    """Edit existing paradigm messages in place, or send new ones if needed."""
    # Try to edit existing messages
    for i, pe in enumerate(para_embeds):
        if i < len(session.paradigm_messages):
            try:
                await session.paradigm_messages[i].edit(embed=pe)
                continue
            except Exception:
                pass
        # Send new message if we don't have one to edit
        msg = await channel.send(embed=pe)
        if i < len(session.paradigm_messages):
            session.paradigm_messages[i] = msg
        else:
            session.paradigm_messages.append(msg)
    # Delete extras if new set has fewer embeds
    while len(session.paradigm_messages) > len(para_embeds):
        try:
            await session.paradigm_messages.pop().delete()
        except Exception:
            pass


async def _resume_from_file(message: discord.Message, session: PrefSession,
                            attachment: discord.Attachment, *, file_bytes: bytes | None = None):
    """Load a progress .xlsx file and resume the session from where it left off."""
    channel = message.channel
    try:
        if file_bytes is None:
            file_bytes = await attachment.read()
        data = load_progress(file_bytes)
        restore_session(session, data)
    except Exception as e:
        await channel.send(embed=discord.Embed(
            title="❌ Failed to Load Progress",
            description=f"Could not read the progress file: {e}",
            color=ERROR_COLOR))
        session.state = "awaiting_csv"
        return

    total_judges = len(session.csv_judges)
    scored = len(session.scores_map)
    embed = discord.Embed(
        title="💾 Progress Restored",
        description=f"Loaded **{total_judges}** judges from `{attachment.filename}`.\n"
                    f"**{scored}** judge(s) already scored.\n"
                    f"Resuming from state: **{session.state}**",
        color=SUCCESS_COLOR)
    await channel.send(embed=embed)

    # Resume into the correct state
    if session.state == "comparing" and session.ranker and not session.ranker.is_complete:
        info_embeds, para_embeds = _build_comparison_embeds(session)
        view = PairwiseView(session)
        await channel.send(
            content=f"▶️ Resuming pairwise — **{session.ranker.remaining}** comparisons left.",
            embeds=info_embeds, view=view)
        await _update_paradigms(channel, session, para_embeds)

    elif session.state == "prompting_scores" and session.current_idx < len(session.unmatched):
        judge = session.unmatched[session.current_idx]
        total = len(session.unmatched)
        idx = session.current_idx + 1
        embed = discord.Embed(title=f"Judge {idx}/{total} — {judge['name']}", color=EMBED_COLOR)
        embed.add_field(name="🏫 School", value=judge.get("school") or "Unknown", inline=True)
        embed.add_field(name="🔄 Rounds", value=str(judge.get("rounds", "?")), inline=True)
        embed.set_footer(text=f"Progress: {idx}/{total}")
        view = ScoringView(session)
        await channel.send(content="▶️ Resuming scoring…", embed=embed, view=view)
        para_embed = _build_paradigm_embed(judge, session)
        await _update_paradigms(channel, session, [para_embed])

    elif session.state == "reviewing":
        review_view = ReviewView(session)
        embed = review_view.build_summary_embed()
        await channel.send(embed=embed, view=review_view)

    elif session.state == "awaiting_quota_mode":
        await send_quota_mode_prompt(channel, session)

    elif session.state == "awaiting_quotas":
        unit = "judges" if session.quota_mode == "judges" else "rounds"
        embed = discord.Embed(
            title="📝 Tier Quotas",
            description=f"Enter **minimum {unit}** for each tier in one line:\n"
                        f"```\nTier1  Tier2  Tier3  Tier4  Tier5  Strike\n```\n"
                        f"Use `-` to skip a tier. Use `min,max` for a range.\n\n"
                        f"**Example:** `6 8 10 8 6 -`",
            color=EMBED_COLOR)
        await channel.send(embed=embed)

    elif session.state == "awaiting_source_choice":
        await _send_source_choice(channel, session)

    elif session.state == "awaiting_unmatched_choice":
        await _send_unmatched_prompt(channel, session)

    elif session.state == "awaiting_rating_range":
        await _send_rating_range_prompt(channel, session)

    elif session.state == "ordinal_bulk_strike":
        await _send_ordinal_bulk_strike(channel, session)

    elif session.state == "ordinal_bucketing":
        await _send_ordinal_bucket(channel, session)

    elif session.state == "ordinal_refine_prompt":
        embed = _build_ordinal_summary_embed(session)
        view = OrdinalRefinePromptView(session)
        await channel.send(embed=embed, view=view)

    elif session.state == "ordinal_refining" and session.ranker and not session.ranker.is_complete:
        tier = session.ordinal_refine_tier or "?"
        info_embeds, para_embeds = _build_comparison_embeds(session)
        view = OrdinalPairwiseView(session)
        await channel.send(
            content=f"▶️ Resuming Tier {tier} refinement — "
                    f"**{session.ranker.remaining}** comparisons left.",
            embeds=info_embeds, view=view)
        await _update_paradigms(channel, session, para_embeds)

    else:
        await channel.send(embed=discord.Embed(
            description=f"Restored to state: **{session.state}**. Send your input to continue.",
            color=EMBED_COLOR))


async def _send_rating_range_prompt(channel, session: PrefSession):
    """Prompt user to choose the tournament's rating scale."""
    embed = discord.Embed(
        title="📏 Rating Scale",
        description="What is the **maximum rating** for this tournament?\n"
                    "(Judges are rated 1 to max; strike = max + 1)\n\n"
                    "Use the buttons below or type a number.",
        color=EMBED_COLOR)
    await channel.send(embed=embed, view=RatingRangeView(session))


async def _apply_rating_range(interaction_or_channel, session: PrefSession, max_val: int):
    """Apply the selected rating range and proceed with pre-filled processing.

    interaction_or_channel: a discord.Interaction (from button) or a channel (from text input).
    If an Interaction, edits the original message; otherwise sends a new one.
    """
    session.rating_max = max_val
    prefilled_count = _process_csv_prefilled(session)
    total = len(session.csv_judges)

    if prefilled_count == total:
        session.matched = []
        session.unmatched = list(session.csv_judges)
        session.state = "awaiting_quota_mode"
        result_embed = discord.Embed(
            description=f"📏 Rating range: **1–{max_val}** (strike = {max_val + 1})\n"
                        f"📋 All **{total}** judges have pre-filled ratings. "
                        f"Proceeding to quota setup…",
            color=SUCCESS_COLOR)
        if isinstance(interaction_or_channel, discord.Interaction):
            await interaction_or_channel.response.edit_message(embed=result_embed, view=None)
            channel = interaction_or_channel.channel
        else:
            channel = interaction_or_channel
            await channel.send(embed=result_embed)
        await send_quota_mode_prompt(channel, session)
    else:
        desc = f"📏 Rating range: **1–{max_val}** (strike = {max_val + 1})"
        if prefilled_count > 0:
            desc += (f"\n📋 **{prefilled_count}** judge(s) have pre-filled ratings "
                     f"(treated as anchors).")
        result_embed = discord.Embed(description=desc, color=EMBED_COLOR)
        if isinstance(interaction_or_channel, discord.Interaction):
            await interaction_or_channel.response.edit_message(embed=result_embed, view=None)
            channel = interaction_or_channel.channel
        else:
            channel = interaction_or_channel
            await channel.send(embed=result_embed)
        session.state = "awaiting_source_choice"
        await _send_source_choice(channel, session)


class SourceChoiceView(ui.View):
    """Buttons for choosing rating source: Notion or from scratch."""
    def __init__(self, session: PrefSession):
        super().__init__(timeout=None)
        self.session = session

    @ui.button(label="Notion Database", style=discord.ButtonStyle.primary, emoji="📚")
    async def notion(self, interaction: discord.Interaction, button: ui.Button):
        await self._handle(interaction, "notion")

    @ui.button(label="From Scratch", style=discord.ButtonStyle.secondary, emoji="📝")
    async def scratch(self, interaction: discord.Interaction, button: ui.Button):
        await self._handle(interaction, "scratch")

    async def _handle(self, interaction: discord.Interaction, choice: str):
        session = self.session
        channel = interaction.channel
        self.stop()

        if choice == "notion":
            await interaction.response.edit_message(
                embed=discord.Embed(description="🔍 Fetching absolute scores from Notion…",
                                    color=EMBED_COLOR), view=None)
            session.notion_judges = fetch_notion_judges()
            session.matched, all_unmatched = match_judges(session.csv_judges, session.notion_judges)
            session.prefilled_unmatched = [j for j in all_unmatched if j["name"] in session.scores_map]
            session.unmatched = [j for j in all_unmatched if j["name"] not in session.scores_map]
            summary = discord.Embed(title="📊 Match Results", color=EMBED_COLOR)
            summary.add_field(name="✅ Matched", value=str(len(session.matched)), inline=True)
            if session.prefilled_unmatched:
                summary.add_field(name="📋 Pre-filled",
                                  value=str(len(session.prefilled_unmatched)), inline=True)
            summary.add_field(name="❓ Unmatched", value=str(len(session.unmatched)), inline=True)
            summary.add_field(name="📚 Notion DB", value=str(len(session.notion_judges)), inline=True)
            # Edit the loading message to show results
            await interaction.edit_original_response(embed=summary)
            if session.unmatched:
                session.state = "awaiting_unmatched_choice"
                await _send_unmatched_prompt(channel, session)
            else:
                session.state = "awaiting_quota_mode"
                await send_quota_mode_prompt(channel, session)

        else:  # scratch
            session.notion_judges = []
            session.matched = []
            all_judges = list(session.csv_judges)
            session.prefilled_unmatched = [j for j in all_judges if j["name"] in session.scores_map]
            session.unmatched = [j for j in all_judges if j["name"] not in session.scores_map]
            if not session.unmatched:
                await interaction.response.edit_message(
                    embed=discord.Embed(
                        description="📋 All judges already have pre-filled ratings. "
                                    "Proceeding to quota setup…",
                        color=SUCCESS_COLOR), view=None)
                session.state = "awaiting_quota_mode"
                await send_quota_mode_prompt(channel, session)
            else:
                desc = f"📝 Ranking from scratch — **{len(session.unmatched)}** judges to rate."
                if session.prefilled_unmatched:
                    desc += (f"\n📋 **{len(session.prefilled_unmatched)}** judge(s) have "
                             f"pre-filled ratings (anchors).")
                await interaction.response.edit_message(
                    embed=discord.Embed(description=desc, color=EMBED_COLOR), view=None)
                session.state = "awaiting_unmatched_choice"
                await _send_unmatched_prompt(channel, session)


class UnmatchedChoiceView(ui.View):
    """Buttons for choosing how to handle unmatched judges."""
    def __init__(self, session: PrefSession):
        super().__init__(timeout=None)
        self.session = session

    @ui.button(label="Direct Rating", style=discord.ButtonStyle.primary, emoji="🎯")
    async def direct_rate(self, interaction: discord.Interaction, button: ui.Button):
        await self._handle(interaction, "rate")

    @ui.button(label="Pairwise Compare", style=discord.ButtonStyle.primary, emoji="🔀")
    async def pairwise(self, interaction: discord.Interaction, button: ui.Button):
        await self._handle(interaction, "compare")

    @ui.button(label="Ordinal Ranking", style=discord.ButtonStyle.primary, emoji="📊")
    async def ordinal(self, interaction: discord.Interaction, button: ui.Button):
        await self._handle(interaction, "ordinal")

    @ui.button(label="Skip", style=discord.ButtonStyle.secondary, emoji="⏭️")
    async def skip(self, interaction: discord.Interaction, button: ui.Button):
        await self._handle(interaction, "skip")

    async def _handle(self, interaction: discord.Interaction, choice: str):
        session = self.session
        channel = interaction.channel
        self.stop()

        if choice == "rate":
            session.state = "prompting_scores"
            session.current_idx = 0
            await interaction.response.edit_message(
                embed=discord.Embed(description="🎯 Starting direct rating…", color=EMBED_COLOR),
                view=None)
            if tabroom_scraper and tabroom_cache:
                loading = await channel.send(embed=discord.Embed(
                    description="🌐 Fetching Tabroom paradigms…", color=EMBED_COLOR))
                for j in session.unmatched:
                    session.paradigms[j["name"]] = tabroom_cache.get_or_fetch(j["name"], tabroom_scraper)
                await loading.delete()
            judge = session.unmatched[0]
            total = len(session.unmatched)
            embed = discord.Embed(title=f"Judge 1/{total} — {judge['name']}", color=EMBED_COLOR)
            embed.add_field(name="🏫 School", value=judge.get("school") or "Unknown", inline=True)
            embed.add_field(name="🔄 Rounds", value=str(judge.get("rounds", "?")), inline=True)
            url = tabroom_paradigm_url(judge["name"])
            embed.add_field(name="🔗 Tabroom", value=f"[View full paradigm]({url})", inline=False)
            embed.set_footer(text=f"Progress: 1/{total}")
            view = ScoringView(session)
            rm = session.rating_max
            intro = discord.Embed(
                title="🎯 Direct Rating",
                description=f"**{total}** judges need scores.\n"
                            f"Use the buttons below to score each judge (1–{rm}).\n"
                            f"• **1** = Best  • **{rm}** = Worst  • Strike/Conflict\n"
                            "• Use **Compare** to view two judges side-by-side\n"
                            "• Use **Previous** to go back and change a score",
                color=EMBED_COLOR)
            await channel.send(embed=intro)
            await channel.send(embed=embed, view=view)
            para_embed = _build_paradigm_embed(judge, session)
            await _update_paradigms(channel, session, [para_embed])

        elif choice == "compare":
            await interaction.response.edit_message(
                embed=discord.Embed(description="🔀 Starting pairwise comparison…", color=EMBED_COLOR),
                view=None)
            if tabroom_scraper and tabroom_cache:
                loading = await channel.send(embed=discord.Embed(
                    description="🌐 Fetching Tabroom paradigms…", color=EMBED_COLOR))
                for j in session.unmatched:
                    session.paradigms[j["name"]] = tabroom_cache.get_or_fetch(j["name"], tabroom_scraper)
                await loading.delete()
            anchors = _select_anchors(session.matched)
            for anchor in anchors:
                prefilled = session.scores_map.get(anchor["name"])
                if prefilled is not None:
                    anchor["score"] = prefilled
            for j in session.prefilled_unmatched:
                score = session.scores_map.get(j["name"])
                if score is not None and score < 6.0:
                    anchors.append({"name": j["name"], "school": j["school"],
                                    "rounds": j["rounds"], "score": score})
            session.ranker = PairwiseRanker(session.unmatched, anchors)
            session.state = "comparing"
            intro = discord.Embed(
                title="🔀 Pairwise Comparison Mode",
                description=f"**{session.ranker.total_comparisons}** matchups to rank "
                            f"**{len(session.unmatched)}** unknown judges.\n"
                            "Pick the better judge using the buttons below.\n"
                            "Press **✅ Done** to finish early.",
                color=EMBED_COLOR)
            await channel.send(embed=intro)
            info_embeds, para_embeds = _build_comparison_embeds(session)
            view = PairwiseView(session)
            await channel.send(embeds=info_embeds, view=view)
            await _update_paradigms(channel, session, para_embeds)

        elif choice == "ordinal":
            session.ordinal_mode = True
            # In ordinal mode, ALL judges need bucketing (matched + unmatched)
            all_to_bucket = list(session.csv_judges)
            session.unmatched = all_to_bucket
            session.current_idx = 0
            await interaction.response.edit_message(
                embed=discord.Embed(description="📊 Starting ordinal ranking…", color=EMBED_COLOR),
                view=None)
            if tabroom_scraper and tabroom_cache:
                loading = await channel.send(embed=discord.Embed(
                    description="🌐 Fetching Tabroom paradigms…", color=EMBED_COLOR))
                for j in session.unmatched:
                    session.paradigms[j["name"]] = tabroom_cache.get_or_fetch(j["name"], tabroom_scraper)
                await loading.delete()
            # Start with bulk strike/conflict
            session.state = "ordinal_bulk_strike"
            await _send_ordinal_bulk_strike(channel, session)

        else:  # skip
            await interaction.response.edit_message(
                embed=discord.Embed(
                    description=f"⏭️ Skipping {len(session.unmatched)} unmatched judge(s) — "
                                "they'll have empty ratings.",
                    color=EMBED_COLOR), view=None)
            session.skipped_judges = list(session.unmatched)
            session.state = "awaiting_quota_mode"
            await send_quota_mode_prompt(channel, session)


async def _send_source_choice(channel, session: PrefSession):
    """Prompt user to choose between Notion database or ranking from scratch."""
    embed = discord.Embed(
        title="📚 Rating Source",
        description="How would you like to rate judges?",
        color=EMBED_COLOR)
    view = SourceChoiceView(session)
    await channel.send(embed=embed, view=view)


async def _send_unmatched_prompt(channel, session: PrefSession):
    embed = discord.Embed(
        title="❓ Judges to Rate",
        description=f"**{len(session.unmatched)}** judge(s) need ratings.\n"
                    "How would you like to handle them?\n\n"
                    "📊 **Ordinal Ranking** — bucket into tiers, then refine within each tier "
                    "via pairwise comparison. Exports a ranked Excel file (no quotas).",
        color=WARN_COLOR)
    view = UnmatchedChoiceView(session)
    await channel.send(embed=embed, view=view)


async def send_quota_mode_prompt(channel, session: PrefSession, interaction=None):
    embed = discord.Embed(title="⚖️ Quota Mode",
                          description="How does this tournament measure tier quotas?", color=EMBED_COLOR)
    embed.add_field(name="Option 1", value="Round count (total available rounds per tier)", inline=False)
    embed.add_field(name="Option 2", value="Judge count (number of judges per tier)", inline=False)
    view = QuotaModeView(session)
    if interaction:
        await interaction.response.edit_message(embed=embed, view=view)
    else:
        await channel.send(embed=embed, view=view)


async def _finish_comparison(channel, session: PrefSession):
    if not session.ranker:
        return

    scores = session.ranker.get_scores()
    rankings = session.ranker.get_rankings()
    for judge, score in scores:
        session.scores_map[judge["name"]] = score

    # Build combined ranked list: directly rated + pairwise ranked, sorted by score
    rated_judges = []  # (name, score, source)
    pairwise_names = set()

    # Pairwise-ranked judges (from Elo)
    score_map = {j["name"]: s for j, s in scores}
    for judge, elo in rankings:
        sc = session.scores_map.get(judge["name"])
        if sc is not None and sc >= 6.0:
            continue  # handled separately
        rated_judges.append((judge["name"], score_map.get(judge["name"], 3.0), f"Elo: {elo:.0f}"))
        pairwise_names.add(judge["name"])

    # Directly rated judges (promoted to anchors during session)
    for anchor in session.ranker.anchors:
        name = anchor["name"]
        sc = anchor.get("score")
        if name in pairwise_names or sc is None:
            continue
        sm_score = session.scores_map.get(name)
        if sm_score is not None and sm_score >= 6.0:
            continue  # handled separately
        # Only include judges that were from this session's CSV (not original Notion anchors)
        if any(j["name"] == name for j in session.csv_judges):
            rated_judges.append((name, sc, "Rated directly"))

    # Sort all rated judges by score (best first)
    rated_judges.sort(key=lambda x: x[1])

    lines = []
    for rank, (name, sc, source) in enumerate(rated_judges, 1):
        lines.append(f"`#{rank}` **{name}** — Score: **{sc}** ({source})")

    # Append struck/conflicted judges
    all_ranked_names = {name for name, _, _ in rated_judges}
    special_lines = []
    for name, sc in session.scores_map.items():
        if name not in all_ranked_names and sc >= 6.0:
            label = _score_label(sc)
            special_lines.append(f"🚫 **{name}** — **{label}**")
    if special_lines:
        lines.append("")
        lines.extend(special_lines)

    total = len(rated_judges) + len(special_lines)
    embed = discord.Embed(
        title=f"📊 Comparison Results ({total} judges: {len(rated_judges)} ranked, {len(special_lines)} struck/conflicted)",
        color=SUCCESS_COLOR)

    # Split across multiple fields to avoid 1024-char limit
    chunk = []
    chunk_len = 0
    field_num = 1
    for line in lines:
        if chunk_len + len(line) + 1 > 1000:
            embed.add_field(name=f"Derived Scores" if field_num == 1 else "\u200b",
                            value="\n".join(chunk), inline=False)
            field_num += 1
            chunk = []
            chunk_len = 0
            # Discord embeds max 25 fields; start new embed if needed
            if field_num > 24:
                await channel.send(embed=embed)
                embed = discord.Embed(color=SUCCESS_COLOR)
                field_num = 1
        chunk.append(line)
        chunk_len += len(line) + 1
    if chunk:
        embed.add_field(name=f"Derived Scores" if field_num == 1 else "\u200b",
                        value="\n".join(chunk), inline=False)
    await channel.send(embed=embed)

    # Clean up paradigm messages
    for msg in session.paradigm_messages:
        try:
            await msg.delete()
        except Exception:
            pass
    session.paradigm_messages.clear()

    # Go straight to quota mode
    session.state = "awaiting_quota_mode"
    await send_quota_mode_prompt(channel, session)


async def run_assignment(channel, session: PrefSession):
    await channel.send(embed=discord.Embed(description="⚙️ Assigning tiers…", color=EMBED_COLOR))

    all_judges = []
    for csv_judge, notion_name, score in session.matched:
        # Pairwise scores_map can override Notion scores if the user compared them
        final_score = session.scores_map.get(csv_judge["name"], score)
        if final_score is None:
            final_score = 4.0
        all_judges.append({
            "name": csv_judge["name"], "school": csv_judge["school"],
            "rounds": csv_judge["rounds"], "score": final_score,
            "notion_name": notion_name})
    for judge in session.unmatched:
        score = session.scores_map.get(judge["name"], 4.0)
        all_judges.append({
            "name": judge["name"], "school": judge["school"],
            "rounds": judge["rounds"], "score": score,
            "notion_name": None})
    for judge in session.prefilled_unmatched:
        score = session.scores_map.get(judge["name"], 4.0)
        all_judges.append({
            "name": judge["name"], "school": judge["school"],
            "rounds": judge["rounds"], "score": score,
            "notion_name": None})

    assigned, report = assign_tiers(all_judges, session.quotas, session.quota_mode)

    report_text = format_report(report)
    report_embed = discord.Embed(title="📊 Tier Assignment Report", color=EMBED_COLOR,
                                 description=f"```\n{report_text}\n```")
    skipped = session.skipped_judges
    if skipped:
        report_embed.add_field(name="ℹ️ Skipped",
                               value=f"{len(skipped)} judge(s) left with empty ratings.", inline=False)
    unmet = [t for t in range(1, 7)
             if t in report and isinstance(report[t], dict) and not report[t].get("met", True)]
    if unmet:
        labels = ["S" if t == 6 else str(t) for t in unmet]
        report_embed.add_field(name="⚠️ Unmet Quotas",
                               value=f"Tier(s): {', '.join(labels)}", inline=False)
    await channel.send(embed=report_embed)

    # Build output CSV
    assigned.sort(key=lambda j: (j["tier"], j["score"]))
    output = io.StringIO()
    writer = csv_mod.writer(output)
    writer.writerow(["First", "Last", "School", "Rounds", "Rating"])
    for judge in assigned:
        tier = judge["tier"]
        rating = "C" if tier == 7 else ("S" if tier == 6 else str(tier))
        first, last = split_name(judge["name"])
        writer.writerow([first, last, judge["school"], judge["rounds"], rating])
    for judge in skipped:
        first, last = split_name(judge["name"])
        writer.writerow([first, last, judge["school"], judge["rounds"], ""])

    output.seek(0)
    file = discord.File(io.BytesIO(output.getvalue().encode("utf-8")), filename="prefs_output.csv")
    await channel.send(embed=discord.Embed(title="✅ Prefs Complete!", color=SUCCESS_COLOR,
                                           description="Here's your filled pref sheet."), file=file)
    session.state = "done"
    sessions.pop(channel.id, None)


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    token = os.environ.get("DISCORD_BOT_TOKEN")
    if not token:
        print("Error: DISCORD_BOT_TOKEN not set in .env")
        sys.exit(1)
    client.run(token)
